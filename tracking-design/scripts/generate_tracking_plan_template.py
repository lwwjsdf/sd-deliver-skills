#!/usr/bin/env python3
"""
generate_tracking_plan_template.py — Generate a Tracking Plan Excel template.

Supports two formats:
  1. standard — Events / Details(Event) / Users  (tracking-design SKILL.md native)
  2. mp      — Custom Event / Preset Event / Public Property / User Attribute
               (tracking_plan.py parser native)

Usage:
    python3 generate_tracking_plan_template.py --format standard --output ./references/tracking-plan.xlsx
    python3 generate_tracking_plan_template.py --format mp --output ./references/tracking-plan-mp.xlsx
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少依赖，请先运行: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_SUBHEADER_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
_SUBHEADER_FONT = Font(bold=True, size=10)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, row_idx: int, cols: int):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def _style_data_cell(cell):
    cell.border = _BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def _auto_width(ws, min_width: int = 12, max_width: int = 50):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                if val_len > max_length:
                    max_length = val_len
            except Exception:
                pass
        adjusted = min(max(min_width, max_length + 2), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ---------------------------------------------------------------------------
# Standard format: Events / Details(Event) / Users
# ---------------------------------------------------------------------------

def _build_standard_events_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Events")
    headers = [
        "Event Variable Name",
        "Event Display Name",
        "Event Attribute Variable Name",
        "Event Attribute Variable Display Name",
        "Date Type",
        "Trigger",
        "Timing",
        "Encryption",
        "Sample Data",
        "Remark",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    # Sample rows
    sample_events = [
        ("$MPLaunch", "小程序启动", "platformType", "平台类型", "String", "MP", "启动时", "N", "Mini Program", "固定值"),
        (None, None, "applicationName", "应用名称", "String", "MP", "启动时", "N", "MyApp", None),
        ("UserLogin", "用户登录", "isSuccess", "是否成功", "Bool", "MP", "登录完成", "N", "true", "true/false"),
        (None, None, "loginMethod", "登录方式", "String", "MP", "登录完成", "N", "手机号登录", "手机号登录；邮箱登录"),
    ]
    for row_idx, row_data in enumerate(sample_events, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data_cell(cell)

    _auto_width(ws)


def _build_standard_details_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Details(Event)")
    headers = [
        "Attribute English variable name",
        "Attribute display name",
        "Data Type",
        "Example",
        "Trigger",
        "Remark",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    sample_props = [
        ("platformType", "平台类型", "String", "Mini Program", "MP", "公共属性"),
        ("applicationName", "应用名称", "String", "MyApp", "MP", "公共属性"),
        ("version", "版本号", "String", "1.0.0", "MP", "公共属性"),
    ]
    for row_idx, row_data in enumerate(sample_props, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data_cell(cell)

    _auto_width(ws)


def _build_standard_users_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Users")
    headers = [
        "Attribute variable name",
        "Attribute display name",
        "Data Type",
        "Example",
        "Encryption",
        "Sample Data",
        "Remark",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    sample_users = [
        ("userId", "用户ID", "String", "U123456", "N", "U123456", "主键"),
        ("registerTime", "注册时间", "DateTime", "2024-01-01T00:00:00", "N", "2024-01-01T00:00:00", None),
        ("membershipLevel", "会员等级", "String", "L1", "N", "L0;L1;L2;L3;L4", "枚举值"),
    ]
    for row_idx, row_data in enumerate(sample_users, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data_cell(cell)

    _auto_width(ws)


def build_standard_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet
    _build_standard_events_sheet(wb)
    _build_standard_details_sheet(wb)
    _build_standard_users_sheet(wb)
    return wb


# ---------------------------------------------------------------------------
# MP format: Custom Event / Preset Event / Public Property / User Attribute
# ---------------------------------------------------------------------------

def _build_mp_custom_event_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Custom Event")
    headers = [
        "Serial",
        "Event English Variable Name",
        "Event Display Name",
        "Event Attribute Variable Name",
        "Event Attribute Display Name",
        "Data Type",
        "Example",
        "Trigger",
        "Timing",
        "Encryption",
        "Remark",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    sample_rows = [
        (1, "UserLogin", "用户登录", "isSuccess", "是否成功", "Bool", "true", "MP", "登录完成", "N", "true/false"),
        (None, None, None, "loginMethod", "登录方式", "String", "手机号登录", "MP", "登录完成", "N", "手机号登录；邮箱登录"),
        (2, "ProductOrder", "商品下单", "productId", "商品ID", "String", "P001", "MP", "下单时", "N", None),
        (None, None, None, "orderAmount", "订单金额", "Number", "199.99", "MP", "下单时", "N", None),
    ]
    for row_idx, row_data in enumerate(sample_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data_cell(cell)

    _auto_width(ws)


def _build_mp_preset_event_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Preset Event")
    headers = [
        "Event English Variable Name",
        "Event Display Name",
        "Event Attribute Variable Name",
        "Event Attribute Display Name",
        "Data Type",
        "Example",
        "Trigger",
        "Remark",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    sample_rows = [
        ("$MPLaunch", "小程序启动", "platformType", "平台类型", "String", "Mini Program", "MP", "SDK预置"),
        (None, None, "applicationName", "应用名称", "String", "MyApp", "MP", "SDK预置"),
        ("$MPShow", "小程序显示", "$url", "页面路径", "String", "/pages/index/index", "MP", "SDK预置"),
    ]
    for row_idx, row_data in enumerate(sample_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data_cell(cell)

    _auto_width(ws)


def _build_mp_public_property_sheet(wb: openpyxl.Workbook):
    # Note: double space in "Public  Property" to match tracking_plan.py parser
    ws = wb.create_sheet("Public  Property")
    headers = [
        "Attribute English variable name",
        "Attribute display name",
        "Data Type",
        "Example",
        "Trigger",
        "Remark",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    sample_rows = [
        ("platformType", "平台类型", "String", "Mini Program", "MP", "所有事件携带"),
        ("applicationName", "应用名称", "String", "MyApp", "MP", "所有事件携带"),
        ("version", "版本号", "String", "1.0.0", "MP", "所有事件携带"),
    ]
    for row_idx, row_data in enumerate(sample_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data_cell(cell)

    _auto_width(ws)


def _build_mp_user_attribute_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("User Attribute")
    headers = [
        "Attribute variable name",
        "Attribute display name",
        "Data Type",
        "Example",
        "Encryption",
        "Remark",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    sample_rows = [
        ("userId", "用户ID", "String", "U123456", "N", "主键"),
        ("registerTime", "注册时间", "DateTime", "2024-01-01T00:00:00", "N", None),
        ("membershipLevel", "会员等级", "String", "L1", "N", "L0;L1;L2;L3;L4"),
    ]
    for row_idx, row_data in enumerate(sample_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data_cell(cell)

    _auto_width(ws)


def build_mp_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _build_mp_custom_event_sheet(wb)
    _build_mp_preset_event_sheet(wb)
    _build_mp_public_property_sheet(wb)
    _build_mp_user_attribute_sheet(wb)
    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="生成埋点方案 Excel 模板")
    parser.add_argument(
        "--format",
        choices=["standard", "mp"],
        default="standard",
        help="模板格式: standard (Events/Details/Users) 或 mp (Custom Event/Preset Event/Public Property/User Attribute)",
    )
    parser.add_argument(
        "--output",
        "-o",
        default="./references/tracking-plan.xlsx",
        help="输出文件路径",
    )
    args = parser.parse_args()

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if args.format == "standard":
        wb = build_standard_workbook()
        print(f"生成标准格式模板: {output_path}")
        print("  Sheets: Events / Details(Event) / Users")
    else:
        wb = build_mp_workbook()
        print(f"生成 MP 格式模板: {output_path}")
        print("  Sheets: Custom Event / Preset Event / Public  Property / User Attribute")

    wb.save(str(output_path))
    print(f"已保存: {output_path}")

    # Optionally update .env
    env_path = output_path.parent.parent / ".env"
    if env_path.exists():
        env_content = env_path.read_text(encoding="utf-8")
        rel_path = os.path.relpath(output_path, env_path.parent)
        if "TRACKING_PLAN_PATH" in env_content:
            lines = env_content.splitlines()
            new_lines = []
            for line in lines:
                if line.startswith("TRACKING_PLAN_PATH="):
                    new_lines.append(f"TRACKING_PLAN_PATH={rel_path}")
                else:
                    new_lines.append(line)
            env_path.write_text("\n".join(new_lines) + "\n", encoding="utf-8")
        else:
            with open(env_path, "a", encoding="utf-8") as f:
                f.write(f"\nTRACKING_PLAN_PATH={rel_path}\n")
        print(f"已更新 .env: TRACKING_PLAN_PATH={rel_path}")


if __name__ == "__main__":
    main()
