"""Tests for validate_pre_import.py."""
import json
from pathlib import Path

import openpyxl
import pytest

from validate_pre_import import PreImportValidator, main


def _make_tracking_plan(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Custom Event")
    ws.append(["No.", "Event Variable Name", "Event Display Name", "Event Attribute Variable Name",
               "Attribute Display Name", "Data Type", "Example", "Trigger", "Timing", "Encryption", "Remarks"])
    ws.append([1, "Login", "登录", "platform", "平台", "string", "MP;Web", "MP", "", "", ""])
    ws.append(["", "", "", "amount", "金额", "number", "", "MP", "", "", ""])
    ws.append([2, "Purchase", "购买", "orderId", "订单ID", "string", "", "MP", "", "", ""])

    ws2 = wb.create_sheet("Public  Property")
    ws2.append(["Attribute English variable name", "Attribute display name", "Data type", "Example", "Trigger", "Remark"])
    ws2.append(["applicationName", "应用名", "string", "App", "MP", ""])

    path = tmp_path / "plan.xlsx"
    wb.save(path)
    return str(path)


def _write_jsonl(tmp_path, records):
    path = tmp_path / "data.jsonl"
    with open(path, "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r) + "\n")
    return str(path)


def test_validate_passes_when_data_matches_plan(tmp_path):
    plan_path = _make_tracking_plan(tmp_path)
    jsonl_path = _write_jsonl(tmp_path, [
        {"type": "track", "event": "Login", "distinct_id": "u1", "time": 1704067200000,
         "properties": {"platform": "MP", "amount": 100, "applicationName": "App"}},
        {"type": "track", "event": "Purchase", "distinct_id": "u1", "time": 1704067200000,
         "properties": {"orderId": "ORD-001", "applicationName": "App"}},
    ])

    from tracking_plan import TrackingPlan
    plan = TrackingPlan(plan_path)
    validator = PreImportValidator(plan)
    report = validator.validate(jsonl_path)

    assert report["passed"] is True
    assert len(report["errors"]) == 0


def test_validate_finds_unknown_event(tmp_path):
    plan_path = _make_tracking_plan(tmp_path)
    jsonl_path = _write_jsonl(tmp_path, [
        {"type": "track", "event": "UnknownEvent", "distinct_id": "u1", "time": 1704067200000,
         "properties": {}},
    ])

    from tracking_plan import TrackingPlan
    plan = TrackingPlan(plan_path)
    validator = PreImportValidator(plan)
    report = validator.validate(jsonl_path)

    assert report["passed"] is False
    assert any(f["rule"] == "event_name" for f in report["errors"])


def test_validate_finds_missing_property(tmp_path):
    plan_path = _make_tracking_plan(tmp_path)
    jsonl_path = _write_jsonl(tmp_path, [
        {"type": "track", "event": "Login", "distinct_id": "u1", "time": 1704067200000,
         "properties": {"applicationName": "App"}},  # missing platform
    ])

    from tracking_plan import TrackingPlan
    plan = TrackingPlan(plan_path)
    validator = PreImportValidator(plan)
    report = validator.validate(jsonl_path)

    assert report["passed"] is False
    assert any(f["rule"] == "missing_property" for f in report["errors"])


def test_validate_finds_type_mismatch(tmp_path):
    plan_path = _make_tracking_plan(tmp_path)
    jsonl_path = _write_jsonl(tmp_path, [
        {"type": "track", "event": "Login", "distinct_id": "u1", "time": 1704067200000,
         "properties": {"platform": "MP", "amount": "not-a-number"}},
    ])

    from tracking_plan import TrackingPlan
    plan = TrackingPlan(plan_path)
    validator = PreImportValidator(plan)
    report = validator.validate(jsonl_path)

    assert any(f["rule"] == "type_mismatch" for f in report["all_findings"])


def test_validate_detects_enum_violation(tmp_path):
    plan_path = _make_tracking_plan(tmp_path)
    jsonl_path = _write_jsonl(tmp_path, [
        {"type": "track", "event": "Login", "distinct_id": "u1", "time": 1704067200000,
         "properties": {"platform": "Invalid", "applicationName": "App"}},
    ])

    from tracking_plan import TrackingPlan
    plan = TrackingPlan(plan_path)
    validator = PreImportValidator(plan)
    report = validator.validate(jsonl_path)

    assert any(f["rule"] == "enum_violation" for f in report["all_findings"])


def test_main_json_output(tmp_path, monkeypatch, capsys):
    plan_path = _make_tracking_plan(tmp_path)
    jsonl_path = _write_jsonl(tmp_path, [
        {"type": "track", "event": "Login", "distinct_id": "u1", "time": 1704067200000,
         "properties": {"platform": "MP", "amount": 100, "applicationName": "App"}},
        {"type": "track", "event": "Purchase", "distinct_id": "u1", "time": 1704067200000,
         "properties": {"orderId": "ORD-001", "applicationName": "App"}},
    ])

    monkeypatch.setenv("TRACKING_PLAN_PATH", plan_path)
    import sys
    sys.argv = ["validate_pre_import.py", "--jsonl", jsonl_path, "--output-json"]

    with pytest.raises(SystemExit) as exc_info:
        main()
    assert exc_info.value.code == 0

    captured = capsys.readouterr()
    assert '"passed": true' in captured.out
