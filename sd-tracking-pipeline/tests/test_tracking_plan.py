"""Tests for tracking_plan.py."""
import datetime

import openpyxl
import pytest

from tracking_plan import (
    TrackingPlan,
    PropertyDef,
    _normalise_type,
    _extract_enum_values,
    _semantic_value,
)


def _make_workbook(tmp_path):
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Custom Event sheet
    ws = wb.create_sheet("Custom Event")
    ws.append(["No.", "Event Variable Name", "Event Display Name", "Event Attribute Variable Name",
               "Attribute Display Name", "Data Type", "Example", "Trigger", "Timing", "Encryption", "Remarks"])
    ws.append([1, "Login", "登录", "platformType", "平台类型", "string", "MP;Web", "MP", "", "", ""])
    ws.append(["", "", "", "amount", "金额", "number", "100", "MP", "", "", ""])
    ws.append([2, "Purchase", "购买", "orderId", "订单ID", "string", "", "MP", "", "", ""])

    # Preset Event sheet
    ws2 = wb.create_sheet("Preset Event")
    ws2.append(["Section"])
    ws2.append(["Event Variable Name", "Event Display Name", "Event Attribute Variable Name",
                "Attribute Display Name", "Attribute Type", "Example", "Trigger/Timing", "Remark"])
    ws2.append(["$MPLaunch", "小程序启动", "$scene", "场景", "string", "1001", "MP", ""])

    # Public Property sheet
    ws3 = wb.create_sheet("Public  Property")
    ws3.append(["Attribute English variable name", "Attribute display name", "Data type", "Example", "Trigger", "Remark"])
    ws3.append(["applicationName", "应用名", "string", "WeChat", "MP", ""])

    # User Attribute sheet
    ws4 = wb.create_sheet("User Attribute")
    ws4.append(["Attribute Variable Name", "Attribute Display Name", "Data Type", "Example"])
    ws4.append(["level", "等级", "string", "vip"])

    path = tmp_path / "tracking_plan.xlsx"
    wb.save(path)
    return str(path)


def test_parse_custom_events(tmp_path):
    path = _make_workbook(tmp_path)
    plan = TrackingPlan(path)
    events = plan.list_events()
    assert "Login" in events
    assert "$MPLaunch" in events

    schema = plan.get_event_schema("Login")
    assert schema is not None
    prop_names = {p.name for p in schema.properties}
    assert "platformType" in prop_names
    assert "amount" in prop_names
    assert "applicationName" in prop_names  # public property merged


def test_parse_preset_events(tmp_path):
    path = _make_workbook(tmp_path)
    plan = TrackingPlan(path)
    schema = plan.get_event_schema("$MPLaunch")
    assert schema is not None
    assert any(p.name == "$scene" for p in schema.properties)


def test_public_and_user_attributes(tmp_path):
    path = _make_workbook(tmp_path)
    plan = TrackingPlan(path)
    public_props = plan.get_public_properties()
    assert any(p.name == "applicationName" for p in public_props)

    user_attrs = plan.get_user_attributes()
    assert any(p.name == "level" for p in user_attrs)


def test_has_mp_events(tmp_path):
    path = _make_workbook(tmp_path)
    plan = TrackingPlan(path)
    assert plan.has_mp_events() is True


def test_generate_value(tmp_path):
    path = _make_workbook(tmp_path)
    plan = TrackingPlan(path)

    bool_prop = PropertyDef(name="isSuccess", value_type="boolean")
    assert plan.generate_value(bool_prop) in (True, False)

    enum_prop = PropertyDef(name="platformType", value_type="string", enum_values=["MP", "Web"])
    assert plan.generate_value(enum_prop) in ("MP", "Web")

    num_prop = PropertyDef(name="amount", value_type="number")
    assert isinstance(plan.generate_value(num_prop), float)

    semantic_prop = PropertyDef(name="orderId", value_type="string")
    val = plan.generate_value(semantic_prop)
    assert isinstance(val, str)
    assert val.startswith("ORD-") or val.startswith("ID-")


def test_normalise_type():
    assert _normalise_type("Boolean") == "boolean"
    assert _normalise_type("Int") == "int"
    assert _normalise_type("list") == "list"
    assert _normalise_type(None) == "string"


def test_extract_enum_values():
    assert _extract_enum_values("A；B；C") == ["A", "B", "C"]
    assert _extract_enum_values('"A"，"B"') == ["A", "B"]
    assert _extract_enum_values("Fixed value: X or Y") == ["X", "Y"]
    assert _extract_enum_values("Some prose text") is None


def test_semantic_value():
    assert isinstance(_semantic_value("orderId"), str)
    assert isinstance(_semantic_value("amount"), float)
    assert isinstance(_semantic_value("email"), str)
    assert "@" in _semantic_value("email")

    base = datetime.datetime(2025, 1, 1)
    calendar = _semantic_value("calendar", base)
    assert isinstance(calendar, str)
    assert "2025" in calendar
