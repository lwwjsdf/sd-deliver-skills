"""Tests for uat_case_parser.py."""
import sys
from pathlib import Path

import pytest

try:
    import openpyxl
except ImportError:
    openpyxl = None

from uat_case_parser import (
    detect_sheet_role,
    guess_automation,
    parse_scenarios,
    parse_xlsx,
    main,
)


pytestmark = pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")


def _make_workbook(tmp_path, sheets):
    """Helper to build an xlsx file from a dict {sheet_name: [[row1], [row2], ...]}."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    path = tmp_path / "uat.xlsx"
    wb.save(path)
    return str(path)


def test_detect_sheet_role_by_title():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Indicators"
    assert detect_sheet_role(ws) == "indicators"

    ws.title = "ID Mapping"
    assert detect_sheet_role(ws) == "id_mapping"

    ws.title = "Permissions"
    assert detect_sheet_role(ws) == "permissions"

    ws.title = "Manual Cases"
    assert detect_sheet_role(ws) == "manual"


def test_detect_sheet_role_by_headers():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["indicator_name", "value"])
    assert detect_sheet_role(ws) == "indicators"

    ws2 = wb.create_sheet()
    ws2.append(["identity", "id_mapping"])
    assert detect_sheet_role(ws2) == "id_mapping"

    ws3 = wb.create_sheet()
    ws3.append(["permission_name", "role"])
    assert detect_sheet_role(ws3) == "permissions"


def test_detect_sheet_role_unknown():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Random"
    ws.append(["foo", "bar"])
    assert detect_sheet_role(ws) == "unknown"


def test_parse_scenarios(tmp_path):
    path = _make_workbook(
        tmp_path,
        {
            "Scenarios": [
                ["Scenario_ID", "Scenario_Name", "Description"],
                ["S1", "Login", "User logs in"],
                ["S2", "Purchase", ""],
                [None, None, None],  # blank row skipped
            ]
        },
    )
    wb = openpyxl.load_workbook(path, data_only=True)
    scenarios = parse_scenarios(wb["Scenarios"])
    assert len(scenarios) == 2
    assert scenarios[0]["scenario_id"] == "S1"
    assert scenarios[0]["scenario_name"] == "Login"
    assert scenarios[1]["scenario_id"] == "S2"


def test_parse_xlsx(tmp_path):
    path = _make_workbook(
        tmp_path,
        {
            "Indicators": [
                ["indicator_name", "expected"],
                ["A", "1"],
                ["B", "2"],
            ],
            "ID Mapping": [
                ["identity", "mapped_id"],
                ["user1", "id1"],
            ],
        },
    )
    result = parse_xlsx(path)
    assert result["meta"]["source"] == "uat.xlsx"
    assert "Indicators" in result["sheets"]
    assert result["sheets"]["Indicators"]["role"] == "indicators"
    assert result["sheets"]["Indicators"]["row_count"] == 2
    assert result["sheets"]["ID Mapping"]["role"] == "id_mapping"


def test_guess_automation():
    assert guess_automation("indicators", {}) == "auto"
    assert guess_automation("id_mapping", {}) == "auto"
    assert guess_automation("permissions", {}) == "manual"
    assert guess_automation("manual", {"test_approach": "manual check"}) == "manual"
    assert guess_automation("manual", {"test_approach": "click button"}) == "manual"
    assert guess_automation("manual", {"test_approach": "sql validation"}) == "auto"


def test_main_exits_on_missing_file(capsys):
    sys.argv = ["uat_case_parser.py", "/nonexistent/file.xlsx"]
    with pytest.raises(SystemExit) as exc_info:
        main()
    assert exc_info.value.code == 1
    captured = capsys.readouterr()
    assert "文件不存在" in captured.out


def test_main_runs_on_valid_file(tmp_path, capsys):
    path = _make_workbook(
        tmp_path,
        {
            "Indicators": [
                ["indicator_name", "expected"],
                ["A", "1"],
            ],
        },
    )
    sys.argv = ["uat_case_parser.py", path]
    main()
    captured = capsys.readouterr()
    assert "解析完毕" in captured.out
    assert "Indicators" in captured.out
