#!/usr/bin/env python3
"""
uat_case_parser.py — 解析 uat-test-case.xlsx，提取 UAT 用例结构

用法：
  python3 uat_case_parser.py <uat-test-case.xlsx> [--output yaml]
"""

import argparse
import sys
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl 库: pip install openpyxl")
    sys.exit(1)


def parse_scenarios(ws) -> list:
    """解析 Scenarios sheet."""
    scenarios = []
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = {str(c or "").strip().lower(): i for i, c in enumerate(row)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        scenario = {}
        for key, idx in headers.items():
            val = row[idx] if idx < len(row) else None
            scenario[key] = str(val).strip() if val else ""
        if scenario.get("scenario_id") or scenario.get("scenario_name"):
            scenarios.append(scenario)
    return scenarios


def detect_sheet_role(ws) -> str:
    """自动检测 sheet 的角色（indicators/id_mapping/permissions）."""
    name = (ws.title or "").lower()
    if "indicat" in name:
        return "indicators"
    if "id" in name and "map" in name:
        return "id_mapping"
    if "permiss" in name:
        return "permissions"
    if "manual" in name:
        return "manual"
    # 检查第一行的列名
    first_row = [str(c or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    row_text = " ".join(first_row)
    if "indicator" in row_text:
        return "indicators"
    if "identity" in row_text or "id_mapping" in row_text:
        return "id_mapping"
    if "permission" in row_text:
        return "permissions"
    return "unknown"


def parse_xlsx(filepath: str) -> dict:
    """解析 uat-test-case.xlsx，返回结构化数据."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    result = {
        "meta": {
            "source": str(Path(filepath).name),
            "sheets": wb.sheetnames,
        },
        "sheets": {}
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        role = detect_sheet_role(ws)

        # 读取所有行
        headers = {}
        rows_data = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = {str(c or "").strip().lower(): j for j, c in enumerate(row)}
                continue
            if all(c is None for c in row):
                continue
            row_dict = {}
            for key, idx in headers.items():
                val = row[idx] if idx < len(row) else None
                row_dict[key] = str(val).strip() if val else ""
            rows_data.append(row_dict)

        result["sheets"][sheet_name] = {
            "role": role,
            "headers": list(headers.keys()),
            "row_count": len(rows_data),
            "rows": rows_data,
        }

    return result


def guess_automation(role: str, row: dict) -> str:
    """猜测用例是否可自动化."""
    if role in ("indicators", "id_mapping"):
        return "auto"
    if role == "permissions":
        return "manual"
    test_approach = str(row.get("test_approach", "")).lower()
    if "manual" in test_approach or "click" in test_approach or "login" in test_approach:
        return "manual"
    return "auto"


def main():
    parser = argparse.ArgumentParser(description="解析 UAT Test Case Excel")
    parser.add_argument("xlsx_file", help="uat-test-case.xlsx 路径")
    parser.add_argument("--output", "-o", default="", help="输出 JSON 文件路径")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx_file):
        print(f"文件不存在: {args.xlsx_file}")
        sys.exit(1)

    result = parse_xlsx(args.xlsx_file)

    # 输出摘要
    print(f"解析完毕: {args.xlsx_file}")
    print(f"  Sheets: {result['meta']['sheets']}")
    for name, sheet in result["sheets"].items():
        print(f"  {name}: {sheet['row_count']} 行, 类型: {sheet['role']}")

    if args.output:
        import json
        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"\n已输出到: {args.output}")


if __name__ == "__main__":
    main()
