#!/usr/bin/env python3
"""
import_to_cdp.py — 从埋点方案 Excel 自动导入元事件和用户属性到神策 CDP

用法：
    python3 tracking-setup-e2e/scripts/import_to_cdp.py

前置条件（在项目根目录的 .env 中配置）：
    SA_HOST             神策 CDP 地址
    SA_PROJECT          项目 ID
    TRACKING_PLAN_PATH  埋点方案 Excel 路径

依赖：
    pip install openpyxl python-dotenv
"""

import json
import sys
import time
from pathlib import Path

# 把项目根目录加入路径，以便引用 shared/
ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ROOT))

try:
    import openpyxl
    from dotenv import load_dotenv
except ImportError:
    print("缺少依赖，请先运行: pip install openpyxl python-dotenv")
    sys.exit(1)

import os
load_dotenv(ROOT / ".env")

from shared.cdp_client import SA_HOST, SA_PROJECT, ensure_session, fetch, validate_env

TRACKING_PLAN_PATH = os.getenv("TRACKING_PLAN_PATH", "")

# ── 数据类型映射 ──────────────────────────────────────────────
TYPE_MAP = {
    "String": "STRING",
    "List": "LIST",
    "Datetime": "DATETIME",
    "Bool": "BOOL",
    "Number": "NUMBER",
}

# 系统保留字段名（无法创建，自动跳过）
RESERVED_FIELD_NAMES = {"Id", "PersonEmail"}


# ── Excel 解析 ────────────────────────────────────────────────
def parse_events(wb: openpyxl.Workbook) -> list[dict]:
    """从 Events sheet 解析元事件列表（第2列=事件名，第3列=显示名）"""
    if "Events" not in wb.sheetnames:
        return []
    ws = wb["Events"]
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not isinstance(row[0], int):
            continue
        original_name = row[1]
        display_name = row[2] if len(row) > 2 else original_name
        if not original_name:
            continue
        events.append({
            "original_name": str(original_name).strip(),
            "display_name": str(display_name or original_name).strip(),
        })
    return events


def parse_event_fields(wb: openpyxl.Workbook, event_name: str) -> list[dict]:
    """从 Details（Event） sheet 解析指定事件的属性列表"""
    sheet_name = "Details（Event） "
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    fields = []
    in_event = False
    for row in ws.iter_rows(values_only=True):
        cell0 = str(row[0]).strip() if row[0] else ""
        if cell0 == event_name:
            in_event = True
            continue
        if not in_event:
            continue
        serial, name, display, dtype = row[0], row[1], row[2], row[3]
        if not isinstance(serial, int):
            if serial:
                break  # 遇到下一个事件的标题行
            continue
        if not name:
            continue
        name = str(name).strip()
        if name in RESERVED_FIELD_NAMES:
            print(f"    跳过保留字段名: {name}")
            continue
        fields.append({
            "display_name": str(display or name).strip(),
            "name": name,
            "data_type": TYPE_MAP.get(str(dtype), "STRING"),
            "field_type": "BASIC",
            "unit_remark": "",
            "remark": "",
        })
    return fields


def parse_user_attrs(wb: openpyxl.Workbook) -> list[dict]:
    """从 Users sheet 解析用户属性列表（第2列=属性名，第3列=显示名，第4列=类型）"""
    if "Users" not in wb.sheetnames:
        return []
    ws = wb["Users"]
    fields = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        serial, name, display, dtype = row[0], row[1], row[2], row[3]
        if not name or not isinstance(serial, int):
            continue
        name = str(name).strip()
        if name in RESERVED_FIELD_NAMES:
            print(f"  跳过保留字段名: {name}")
            continue
        fields.append({
            "display_name": str(display or name).strip(),
            "name": name,
            "data_type": TYPE_MAP.get(str(dtype), "STRING"),
            "field_type": "BASIC",
            "unit_remark": "",
            "remark": "",
        })
    return fields


# ── API 操作 ──────────────────────────────────────────────────
def create_meta_event(event: dict, fields: list[dict]) -> bool:
    """创建元事件并写入自定义属性"""
    name = event["original_name"]

    # Step 1: 创建事件（带 appKey 基础字段）
    resp = fetch(
        f"/api/v2/horizon/v1/web/event_schema/create?event_type=META&project={SA_PROJECT}",
        method="POST",
        body={
            "event_type": "META",
            "entity_name": "user",
            "meta": {
                "physical_name": "events",
                "entity_name": "user",
                "event_type": "META",
                "original_name": name,
                "display_name": event["display_name"],
                "need_track_platforms": [{"name": "MINI_APP"}],
                "fields": [{
                    "display_name": "appKey", "name": "appKey",
                    "data_type": "STRING", "field_type": "BASIC",
                    "visible": False, "unit_remark": "", "remark": "",
                }],
            },
        },
    )
    if resp.get("code") != "SUCCESS":
        print(f"  ✗ 创建失败: {str(resp.get('message') or resp.get('error', ''))[:100]}")
        return False

    if not fields:
        print(f"  ✓ 创建成功（无自定义属性）")
        return True

    # Step 2: 获取现有字段（保留虚拟字段，排除 preset/builtin）
    detail = fetch(
        f"/api/v2/horizon/v1/web/event_schema/detail"
        f"?event_type=META&project={SA_PROJECT}&name=events.{name}&entity_name=user"
    )
    existing = detail.get("data", {}).get("meta", {}).get("fields", [])
    keep = [f for f in existing if not f.get("is_preset") and not f.get("is_builtin")]

    # Step 3: update 补充自定义属性
    resp2 = fetch(
        f"/api/v2/horizon/v1/web/event_schema/update"
        f"?event_type=META&project={SA_PROJECT}&name=events.{name}&entity_name=user",
        method="POST",
        body={
            "event_type": "META",
            "entity_name": "user",
            "meta": {
                "physical_name": "events",
                "entity_name": "user",
                "event_type": "META",
                "name": f"events.{name}",
                "original_name": name,
                "display_name": event["display_name"],
                "need_track_platforms": [{"name": "MINI_APP"}],
                "fields": keep + fields,
            },
        },
    )
    if resp2.get("code") == "SUCCESS":
        print(f"  ✓ 创建成功，写入 {len(fields)} 个属性")
        return True
    else:
        print(f"  ✗ 属性写入失败: {str(resp2.get('message', ''))[:100]}")
        return False


def insert_user_attrs(fields: list[dict]) -> dict:
    """逐个插入用户属性（避免单个失败影响全部）"""
    ok, skipped, failed = [], [], []
    for f in fields:
        resp = fetch(
            f"/api/v2/horizon/v1/web/schema/field/batch_insert?project={SA_PROJECT}",
            method="POST",
            body={"fields": [f], "schema_class": "USER", "physical_name": "users"},
        )
        code = resp.get("code", "")
        msg = str(resp.get("message") or resp.get("error", ""))
        if code == "SUCCESS":
            ok.append(f["name"])
        elif "ALREADY_EXISTS" in msg or "already exist" in msg.lower():
            skipped.append(f["name"])
        else:
            failed.append({"name": f["name"], "error": msg[:80]})
        time.sleep(0.1)

    print(f"  新增: {len(ok)}  已存在: {len(skipped)}  失败: {len(failed)}")
    for f in failed:
        print(f"    ✗ {f['name']}: {f['error']}")
    return {"ok": ok, "skipped": skipped, "failed": failed}


# ── 主流程 ────────────────────────────────────────────────────
def main():
    validate_env(["SA_HOST", "SA_PROJECT", "TRACKING_PLAN_PATH"])

    excel_path = Path(TRACKING_PLAN_PATH)
    if not excel_path.exists():
        print(f"错误：找不到埋点方案文件：{excel_path}")
        sys.exit(1)

    print(f"=== 神策 CDP 元数据导入 ===")
    print(f"文件: {excel_path.name}")
    print(f"目标: {SA_HOST}  项目: {SA_PROJECT}\n")

    ensure_session()

    wb = openpyxl.load_workbook(excel_path)
    results = {"events": [], "user_attrs": {}}

    # ── 元事件 ──
    print("── 元事件导入 ──")
    events = parse_events(wb)
    if not events:
        print("  未找到 Events sheet 或无数据，跳过")
    else:
        print(f"  发现 {len(events)} 个事件")
        for event in events:
            print(f"\n  [{event['original_name']}]")
            fields = parse_event_fields(wb, event["original_name"])
            success = create_meta_event(event, fields)
            results["events"].append({
                "name": event["original_name"],
                "success": success,
                "fields": len(fields),
            })

    # ── 用户属性 ──
    print("\n── 用户属性导入 ──")
    user_fields = parse_user_attrs(wb)
    if not user_fields:
        print("  未找到 Users sheet 或无数据，跳过")
    else:
        print(f"  发现 {len(user_fields)} 个属性")
        results["user_attrs"] = insert_user_attrs(user_fields)

    # ── 汇总 ──
    print("\n=== 导入完成 ===")
    if results["events"]:
        ok_n = sum(1 for e in results["events"] if e["success"])
        print(f"元事件: {ok_n}/{len(results['events'])} 成功")
    if results["user_attrs"]:
        ua = results["user_attrs"]
        print(f"用户属性: 新增 {len(ua.get('ok', []))}，"
              f"已存在 {len(ua.get('skipped', []))}，"
              f"失败 {len(ua.get('failed', []))}")


if __name__ == "__main__":
    main()
