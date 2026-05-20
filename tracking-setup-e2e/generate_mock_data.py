#!/usr/bin/env python3
"""
Generate mock event data from a Sensors Data tracking plan Excel file.
Outputs JSON files ready for Sensors Data HTTP Track API or batch import.

Usage:
    python3 generate_mock_data.py --input "Tracking Plan.xlsx" --output ./mock_data
    python3 generate_mock_data.py --input "Tracking Plan.xlsx" --output ./mock_data --count 100
"""

import argparse
import base64
import json
import os
import random
import string
import time
import uuid
from datetime import datetime, timedelta

import openpyxl


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def parse_tracking_plan(xlsx_path: str) -> dict:
    """Parse a Sensors Data tracking plan Excel into a structured dict."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    plan = {"events": [], "users": []}

    # --- Events sheet ---
    if "Events" in wb.sheetnames:
        ws = wb["Events"]
        rows = list(ws.iter_rows(values_only=True))

        # Find header row (contains 'Event Variable Name')
        header_row_idx = None
        for i, row in enumerate(rows):
            if any(cell == "Event Variable Name" for cell in row if cell):
                header_row_idx = i
                break

        if header_row_idx is not None:
            headers = rows[header_row_idx]
            col = {v: i for i, v in enumerate(headers) if v}

            current_event = None
            for row in rows[header_row_idx + 1:]:
                event_name = row[col.get("Event Variable Name", 1)] if col.get("Event Variable Name") is not None else None
                attr_name = row[col.get("Event  Attribute Variable Name", 3)] if col.get("Event  Attribute Variable Name") is not None else None
                attr_display = row[col.get("Event  Attribute Variable Display Name", 4)] if col.get("Event  Attribute Variable Display Name") is not None else None
                data_type = row[col.get("Date Type", 5)] if col.get("Date Type") is not None else None
                sample = row[col.get("Sample Data", 11)] if col.get("Sample Data") is not None else None

                if event_name:
                    current_event = {
                        "name": event_name,
                        "display_name": row[col.get("Event Display Name", 2)] if col.get("Event Display Name") is not None else event_name,
                        "properties": []
                    }
                    plan["events"].append(current_event)

                if current_event and attr_name:
                    current_event["properties"].append({
                        "name": attr_name,
                        "display_name": attr_display or attr_name,
                        "type": data_type or "String",
                        "sample": sample,
                    })

    # --- Users sheet ---
    if "Users" in wb.sheetnames:
        ws = wb["Users"]
        rows = list(ws.iter_rows(values_only=True))

        header_row_idx = None
        for i, row in enumerate(rows):
            if any(cell == "Attribute variable name" for cell in row if cell):
                header_row_idx = i
                break

        if header_row_idx is not None:
            headers = rows[header_row_idx]
            col = {v: i for i, v in enumerate(headers) if v}

            for row in rows[header_row_idx + 1:]:
                attr_name = row[col.get("Attribute variable name", 1)] if col.get("Attribute variable name") is not None else None
                if attr_name:
                    plan["users"].append({
                        "name": attr_name,
                        "display_name": row[col.get("Attribute display name", 2)] if col.get("Attribute display name") is not None else attr_name,
                        "type": row[col.get("Data Type", 3)] if col.get("Data Type") is not None else "String",
                        "sample": row[col.get("Sample Data", 9)] if col.get("Sample Data") is not None else None,
                    })

    return plan


# ---------------------------------------------------------------------------
# Mock value generation
# ---------------------------------------------------------------------------

def _random_string(length=8):
    return "".join(random.choices(string.ascii_letters + string.digits, k=length))


def _random_datetime(days_back=90):
    delta = timedelta(days=random.randint(0, days_back), seconds=random.randint(0, 86400))
    return (datetime.now() - delta).strftime("%Y-%m-%dT%H:%M:%S.000Z")


def _random_timestamp_ms(days_back=90):
    delta = timedelta(days=random.randint(0, days_back), seconds=random.randint(0, 86400))
    return int((datetime.now() - delta).timestamp() * 1000)


def generate_value(prop: dict) -> object:
    """Generate a realistic mock value for a property based on its type and sample."""
    dtype = (prop.get("type") or "String").strip().lower()
    sample = prop.get("sample")

    # Use sample value directly when it's simple and concrete
    if sample is not None and not isinstance(sample, bool):
        sample_str = str(sample).strip()
        # If sample contains semicolons, it's a picklist — pick one
        if ";" in sample_str:
            return random.choice([s.strip() for s in sample_str.split(";") if s.strip()])
        # Use sample as-is for short concrete values
        if len(sample_str) < 60 and "\n" not in sample_str:
            return sample

    if dtype in ("bool", "boolean", "checkbox"):
        return random.choice([True, False])

    if dtype in ("datetime", "date/time", "date"):
        return _random_datetime()

    if dtype in ("number", "int", "integer", "double", "currency"):
        return round(random.uniform(1, 10000), 2)

    if dtype in ("list", "picklist", "picklist (multi-select)", "set"):
        options = ["Option_A", "Option_B", "Option_C", "Option_D"]
        return random.choice(options)

    # Default: String
    return _random_string(12)


def generate_user_id() -> str:
    return f"user_{uuid.uuid4().hex[:12]}"


# ---------------------------------------------------------------------------
# Sensors Data event record builder
# ---------------------------------------------------------------------------

def build_track_record(event: dict, distinct_id: str, event_time_ms: int) -> dict:
    properties = {
        "$app_version": "1.0.0",
        "$lib": "python",
        "$lib_version": "1.0.0",
    }
    for prop in event.get("properties", []):
        properties[prop["name"]] = generate_value(prop)

    return {
        "distinct_id": distinct_id,
        "type": "track",
        "event": event["name"],
        "time": event_time_ms,
        "properties": properties,
    }


def build_profile_record(user_attrs: list, distinct_id: str) -> dict:
    properties = {}
    for attr in user_attrs:
        properties[attr["name"]] = generate_value(attr)
    return {
        "distinct_id": distinct_id,
        "type": "profile_set",
        "time": int(time.time() * 1000),
        "properties": properties,
    }


# ---------------------------------------------------------------------------
# Output formats
# ---------------------------------------------------------------------------

def to_batch_payload(records: list) -> str:
    """Encode records as base64 for Sensors Data batch HTTP API."""
    data = json.dumps(records, ensure_ascii=False)
    return base64.b64encode(data.encode("utf-8")).decode("utf-8")


class _JSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.strftime("%Y-%m-%dT%H:%M:%S.000Z")
        if hasattr(obj, "isoformat"):
            return obj.isoformat()
        if hasattr(obj, "__str__"):
            return str(obj)
        return super().default(obj)


def write_outputs(records: list, output_dir: str, prefix: str):
    os.makedirs(output_dir, exist_ok=True)

    # Raw JSON (one record per line, for readability and streaming import)
    jsonl_path = os.path.join(output_dir, f"{prefix}.jsonl")
    with open(jsonl_path, "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False, cls=_JSONEncoder) + "\n")

    # Batch payload (base64, ready to POST to /sa?token=xxx)
    batch_path = os.path.join(output_dir, f"{prefix}_batch.txt")
    with open(batch_path, "w", encoding="utf-8") as f:
        data = json.dumps(records, ensure_ascii=False, cls=_JSONEncoder)
        f.write(base64.b64encode(data.encode("utf-8")).decode("utf-8"))

    return jsonl_path, batch_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate mock data from Sensors Data tracking plan")
    parser.add_argument("--input", required=True, help="Path to tracking plan Excel file")
    parser.add_argument("--output", default="./mock_data", help="Output directory")
    parser.add_argument("--count", type=int, default=50, help="Records per event (default: 50)")
    parser.add_argument("--users", type=int, default=20, help="Number of simulated users (default: 20)")
    args = parser.parse_args()

    print(f"Parsing tracking plan: {args.input}")
    plan = parse_tracking_plan(args.input)

    events = plan["events"]
    user_attrs = plan["users"]

    print(f"Found {len(events)} events, {len(user_attrs)} user attributes")

    # Generate a pool of user IDs
    user_ids = [generate_user_id() for _ in range(args.users)]

    all_records = []

    # User profile records
    if user_attrs:
        for uid in user_ids:
            all_records.append(build_profile_record(user_attrs, uid))

    # Event records
    for event in events:
        for _ in range(args.count):
            uid = random.choice(user_ids)
            ts = _random_timestamp_ms()
            all_records.append(build_track_record(event, uid, ts))

    # Shuffle to simulate realistic interleaving
    random.shuffle(all_records)

    jsonl_path, batch_path = write_outputs(all_records, args.output, "mock_events")

    print(f"\nGenerated {len(all_records)} records")
    print(f"  JSONL:       {jsonl_path}")
    print(f"  Batch (b64): {batch_path}")
    print(f"\nTo import via HTTP API:")
    print(f"  curl -X POST 'https://<your-sa-host>/sa?token=<token>' \\")
    print(f"       -d 'data='$(cat {batch_path})")


if __name__ == "__main__":
    main()
