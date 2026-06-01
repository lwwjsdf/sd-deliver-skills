"""
tracking_plan.py — Parse the Tracking Plan Excel and expose event/property schemas.

Supports sheets: Custom Event, Preset Event, Public Property (with double space), User Attribute.
"""

from __future__ import annotations

import random
import re
import string
from dataclasses import dataclass, field
from typing import Dict, List, Optional

import openpyxl


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass
class PropertyDef:
    name: str
    required: bool = False
    value_type: str = "string"  # string / boolean / number / list / bool / datetime
    trigger: str = "MP"  # MP / Server / Web / Mini program
    enum_values: Optional[List[str]] = None  # non-empty when enum constraint exists
    description: str = ""


@dataclass
class EventSchema:
    event_name: str
    trigger: str = "MP"
    properties: List[PropertyDef] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _normalise_type(raw: Optional[str]) -> str:
    """Map Excel type strings to canonical value_type tokens."""
    if not raw:
        return "string"
    t = raw.strip().lower()
    if t in ("bool", "boolean"):
        return "boolean"
    if t in ("number", "int", "integer", "float", "double", "datetime"):
        return t
    if t == "list":
        return "list"
    return "string"


def _extract_enum_values(text: Optional[str]) -> Optional[List[str]]:
    """
    Try to extract a finite set of enum values from an example/description cell.

    Patterns recognised:
      - Values separated by Chinese/ASCII semicolons: "A；B；C"
      - Quoted tokens: "手机号登录"，"邮箱登录"
      - "Fixed value: X" or "Fixed value: X or Y"
    Returns None when no clear enumeration is found.
    """
    if not text:
        return None

    text = text.strip()

    # Fixed value pattern: "Fixed value: X" or "Fixed value: X or Y"
    fixed_match = re.match(r"[Ff]ixed\s+value[：:]\s*(.+)", text)
    if fixed_match:
        raw = fixed_match.group(1)
        # Split on " or " and strip quotes
        parts = re.split(r"\s+or\s+", raw)
        values = [p.strip().strip('"').strip('"').strip('"').strip() for p in parts]
        values = [v for v in values if v]
        if values:
            return values

    # Chinese-semicolon separated list
    if "；" in text:
        parts = text.split("；")
        values = [p.strip().strip('"').strip('"').strip('"').strip() for p in parts]
        values = [v for v in values if v and len(v) < 60]
        if 2 <= len(values) <= 20:
            return values

    # ASCII semicolon separated (but not prose sentences)
    if ";" in text and text.count(";") >= 1:
        parts = text.split(";")
        values = [p.strip().strip('"').strip('"').strip('"').strip() for p in parts]
        values = [v for v in values if v and len(v) < 60 and " " not in v.strip()]
        if 2 <= len(values) <= 20:
            return values

    # Quoted tokens pattern: "A"，"B"，"C"
    quoted = re.findall(r'["""]([^"""]{1,50})["""]', text)
    if len(quoted) >= 2:
        return quoted

    return None


def _rand_string(length: int = 8) -> str:
    return "".join(random.choices(string.ascii_letters + string.digits, k=length))


# ---------------------------------------------------------------------------
# Semantic value generator — infers format from field name
#
# LAYERING PRINCIPLE:
#   tracking_plan.py  → universal format rules (IDs, amounts, dates, contacts)
#                        applies to any project without modification
#   business_logic.yaml property_enums → industry/project-specific content
#                        (product names, search keywords, app name, login methods)
# ---------------------------------------------------------------------------

import datetime as _dt

# Patterns: (regex on field name, generator callable)
# Checked in order; first match wins.
_SEMANTIC_PATTERNS: list[tuple[re.Pattern, object]] = []


def _reg(pattern: str, fn):
    _SEMANTIC_PATTERNS.append((re.compile(pattern, re.IGNORECASE), fn))


# ── ID / Code formats (universal) ────────────────────────────────────────────
_reg(r"orderid|refundid|ordernum",
     lambda: f"ORD-{random.randint(100000, 999999)}")
_reg(r"productid|productcode",
     lambda: f"P{random.randint(10000, 99999):05d}")
_reg(r"ticketid(?!list)|ticketid$",
     lambda: f"TK-{random.randint(100000, 999999)}")
_reg(r"membershipcardid|membershipnumber",
     lambda: f"M-{_dt.date.today().year}-{random.randint(100000, 999999):06d}")
_reg(r"voucherid(?!list)|voucherid$",
     lambda: f"VC-{random.randint(10000, 99999)}")
_reg(r"\bid\b|_id$|id$",
     lambda: f"ID-{random.randint(10000, 99999)}")
_reg(r"code$",
     lambda: f"C{random.randint(1000, 9999)}")

# ── Page / URL (universal format, project overrides paths via property_enums) ─
_reg(r"pagename|landingpagename",
     lambda: random.choice(["首页", "列表页", "详情页", "搜索结果页", "个人中心"]))
_reg(r"url|pageurl|landingpageurl",
     lambda: random.choice([
         "/pages/index/index", "/pages/list/index",
         "/pages/detail/index", "/pages/search/result",
         "/pages/profile/index",
     ]))

# ── Amounts (universal ranges) ────────────────────────────────────────────────
_reg(r"discountamount|promotionamount",
     lambda: round(random.uniform(0, 200), 2))
_reg(r"paidamount|refundamount",
     lambda: round(random.uniform(100, 2000), 2))
_reg(r"price|amount|value",
     lambda: round(random.uniform(50, 3000), 2))

# ── Quantities / rankings (universal) ────────────────────────────────────────
_reg(r"quantity|numer|number(?!$)",
     lambda: random.randint(1, 5))
_reg(r"sort|rank|depth",
     lambda: random.randint(1, 20))
_reg(r"duration",
     lambda: round(random.uniform(5, 300), 1))

# ── Dates (universal formats, anchored to event timestamp) ───────────────────
# These lambdas accept an optional base_date; _semantic_value injects it.
_reg(r"calendar",
     lambda base=None: (
         (base or _dt.date.today()) + _dt.timedelta(days=random.randint(1, 180))
     ).strftime("%Y.%m.%d"))
_reg(r"expirationdate|expiry",
     lambda base=None: (
         _dt.datetime.combine(base or _dt.date.today(), _dt.time()) +
         _dt.timedelta(days=random.randint(30, 730))
     ).strftime("%Y-%m-%dT%H:%M:%S"))

# ── Contact / identity (universal formats) ───────────────────────────────────
_reg(r"email",
     lambda: f"user{random.randint(100, 999)}@example.com")
_reg(r"mobile|phone",
     lambda: f"+86 138-{random.randint(1000,9999)}-{random.randint(1000,9999)}")
_reg(r"residence",
     lambda: random.choice(["中国 香港", "中国 北京市", "中国 上海市", "中国 广州市", "海外"]))
_reg(r"transferperson",
     lambda: f"1{random.randint(30,99)}****{random.randint(1000,9999)}")

# ── UTM (universal channel names) ────────────────────────────────────────────
_reg(r"utm_source",
     lambda: random.choice(["wechat", "weibo", "xiaohongshu", "douyin", "direct"]))
_reg(r"utm_medium",
     lambda: random.choice(["social", "cpc", "email", "organic"]))

# ── Version (universal format) ───────────────────────────────────────────────
_reg(r"^version$",
     lambda: random.choice(["1.0.0", "1.0.1", "1.1.0"]))

# ── List fields fallback (universal) ─────────────────────────────────────────
# Project-specific list content should be set via property_enums in business_logic.yaml
_reg(r"idlist$|typelist$|namelist$",
     lambda: _rand_string(8))


def _semantic_value(prop_name: str, ref_dt: "_dt.datetime | None" = None) -> "object | None":
    """Return a semantically appropriate value based on field name, or None if no pattern matches.

    ref_dt: the event timestamp to use as reference for date/time generation.
            Defaults to now() when None.
    """
    name = prop_name.lstrip("$")
    base = ref_dt or _dt.datetime.now()
    base_date = base.date()

    for pattern, fn in _SEMANTIC_PATTERNS:
        if pattern.search(name):
            # Inject ref date into date-sensitive lambdas
            if pattern.pattern in (r"calendar", r"expirationdate|expiry"):
                return fn(base_date)
            return fn()
    return None


# ---------------------------------------------------------------------------
# TrackingPlan
# ---------------------------------------------------------------------------


class TrackingPlan:
    """
    Parse the Tracking Plan Excel workbook and expose event/property schemas.

    Sheets consumed:
      - "Custom Event"      — custom events with per-row properties
      - "Preset Event"      — SDK preset events
      - "Public  Property"  — public properties attached to every event
      - "User Attribute"    — user profile attributes
    """

    def __init__(self, excel_path: str) -> None:
        self._custom_events: Dict[str, EventSchema] = {}
        self._preset_events: Dict[str, EventSchema] = {}
        self._public_props: List[PropertyDef] = []
        self._user_attrs: List[PropertyDef] = []

        wb = openpyxl.load_workbook(excel_path, data_only=True)

        self._parse_public_property(wb)
        self._parse_custom_events(wb)
        self._parse_preset_events(wb)
        self._parse_user_attributes(wb)

    # ------------------------------------------------------------------
    # Sheet parsers
    # ------------------------------------------------------------------

    # Sheet name mappings: canonical -> list of possible names
    SHEET_MAPPINGS = {
        "custom_event": ["custom event", "events", "event"],
        "preset_event": ["preset event", "edm behavior preset events", "preset events"],
        "public_property": [
            "public property",
            "public  property",
            "details（event）",
            "details",
        ],
        "user_attribute": ["user attribute", "users", "user traits", "user"],
    }

    def _find_sheet(self, wb: openpyxl.Workbook, sheet_type: str) -> Optional[str]:
        """Find sheet by canonical type, supporting multiple naming conventions."""
        candidates = self.SHEET_MAPPINGS.get(sheet_type, [sheet_type])
        for name in wb.sheetnames:
            normalized = re.sub(r"[\s（）()]+", " ", name.strip().lower()).strip()
            for candidate in candidates:
                if normalized == candidate or normalized.startswith(candidate):
                    return name
        return None

    def _parse_public_property(self, wb: openpyxl.Workbook) -> None:
        sheet_name = self._find_sheet(wb, "public_property")
        if sheet_name is None:
            return

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find the header row (contains "Attribute English variable name")
        header_idx = None
        for i, row in enumerate(rows):
            if (
                row
                and row[0]
                and "attribute english variable name" in str(row[0]).lower()
            ):
                header_idx = i
                break
        if header_idx is None:
            return

        for row in rows[header_idx + 1 :]:
            attr_name = row[0] if row[0] else None
            if not attr_name:
                continue
            attr_name = str(attr_name).strip()
            if not attr_name:
                continue

            data_type = _normalise_type(str(row[3]) if row[3] else None)
            example = str(row[4]) if row[4] else ""
            trigger = str(row[5]).strip() if row[5] else "MP"

            prop = PropertyDef(
                name=attr_name,
                required=False,
                value_type=data_type,
                trigger=trigger,
                enum_values=_extract_enum_values(example),
                description=example,
            )
            self._public_props.append(prop)

    def _parse_custom_events(self, wb: openpyxl.Workbook) -> None:
        sheet_name = self._find_sheet(wb, "custom_event")
        if sheet_name is None:
            return

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Row 0 is the header; data starts at row 1
        # Columns: 0=serial, 1=event_name, 2=event_display, 3=attr_name,
        #          4=attr_cn, 5=data_type, 6=example, 7=trigger, 8=timing,
        #          9=encryption, 10=remarks
        current_event: Optional[str] = None
        current_trigger: str = "MP"

        for row in rows[1:]:
            # A new event starts when column 1 (event name) is non-empty
            if row[1]:
                current_event = str(row[1]).strip()
                current_trigger = str(row[7]).strip() if row[7] else "MP"
                if current_event not in self._custom_events:
                    self._custom_events[current_event] = EventSchema(
                        event_name=current_event,
                        trigger=current_trigger,
                    )

            if current_event is None:
                continue

            # Attribute name is in column 3
            attr_name = row[3]
            if not attr_name:
                continue
            attr_name = str(attr_name).strip()
            # Some cells have newlines (multiple aliases); take the first token
            attr_name = attr_name.split("\n")[0].strip()
            if not attr_name:
                continue

            data_type = _normalise_type(str(row[5]) if row[5] else None)
            example = str(row[6]) if row[6] else ""

            prop = PropertyDef(
                name=attr_name,
                required=False,
                value_type=data_type,
                trigger=current_trigger,
                enum_values=_extract_enum_values(example),
                description=example,
            )
            self._custom_events[current_event].properties.append(prop)

    def _parse_preset_events(self, wb: openpyxl.Workbook) -> None:
        sheet_name = self._find_sheet(wb, "preset_event")
        if sheet_name is None:
            return

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Row 0: section title, Row 1: header, data from row 2
        # Columns: 0=event_name, 1=event_display, 2=attr_name, 3=attr_display,
        #          4=attr_type, 5=example, 6=trigger_timing, 7=remark,
        #          8=encryption, 9=remarks
        current_event: Optional[str] = None
        current_trigger: str = "MP"

        for row in rows[2:]:
            if row[0]:
                current_event = str(row[0]).strip()
                current_trigger = str(row[6]).strip() if row[6] else "MP"
                if current_event not in self._preset_events:
                    self._preset_events[current_event] = EventSchema(
                        event_name=current_event,
                        trigger=current_trigger,
                    )

            if current_event is None:
                continue

            attr_name = row[2]
            if not attr_name:
                continue
            attr_name = str(attr_name).strip()
            if not attr_name:
                continue

            data_type = _normalise_type(str(row[4]) if row[4] else None)
            example = str(row[5]) if row[5] else ""

            prop = PropertyDef(
                name=attr_name,
                required=False,
                value_type=data_type,
                trigger=current_trigger,
                enum_values=_extract_enum_values(example),
                description=example,
            )
            self._preset_events[current_event].properties.append(prop)

    def _parse_user_attributes(self, wb: openpyxl.Workbook) -> None:
        sheet_name = self._find_sheet(wb, "user_attribute")
        if sheet_name is None:
            return

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Row 0 is the header
        # Columns: 0=attr_name, 1=attr_cn, 2=data_type, 3=example, 4=encryption, 5=remarks
        for row in rows[1:]:
            attr_name = row[0]
            if not attr_name:
                continue
            attr_name = str(attr_name).strip()
            if not attr_name:
                continue

            data_type = _normalise_type(str(row[2]) if row[2] else None)
            example = str(row[3]) if row[3] else ""

            prop = PropertyDef(
                name=attr_name,
                required=False,
                value_type=data_type,
                trigger="MP",
                enum_values=_extract_enum_values(example),
                description=example,
            )
            self._user_attrs.append(prop)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def get_event_schema(self, event_name: str) -> Optional[EventSchema]:
        """Return the full event schema (event props + public props), or None."""
        schema = self._custom_events.get(event_name) or self._preset_events.get(
            event_name
        )
        if schema is None:
            return None

        # Merge public properties (avoid duplicates by name)
        existing_names = {p.name for p in schema.properties}
        merged_props = list(schema.properties)
        for pub_prop in self._public_props:
            if pub_prop.name not in existing_names:
                merged_props.append(pub_prop)

        return EventSchema(
            event_name=schema.event_name,
            trigger=schema.trigger,
            properties=merged_props,
        )

    def get_public_properties(self) -> List[PropertyDef]:
        """Return the list of public properties (platformType, applicationName, version, …)."""
        return list(self._public_props)

    def get_user_attributes(self) -> List[PropertyDef]:
        """Return user attribute definitions from the User Attribute sheet."""
        return list(self._user_attrs)

    def list_events(self) -> List[str]:
        """Return all defined event names (custom + preset)."""
        return sorted(
            list(self._custom_events.keys()) + list(self._preset_events.keys())
        )

    def has_mp_events(self) -> bool:
        """Check if the tracking plan contains any Mini Program preset events ($MP*)."""
        return any(name.startswith("$") for name in self.list_events())

    def generate_value(self, prop: PropertyDef, event_ts: "_dt.datetime | None" = None) -> object:
        """
        Generate a plausible mock value for a property.

        event_ts: event timestamp used as reference for date/time generation.
                  Defaults to now() when None.

        Rules (in priority order):
          1. enum_values present → random.choice(enum_values)
          2. value_type == boolean/bool → random.choice([True, False])
          3. value_type == list → random.choice(enum_values) if enum_values else _rand_string
          4. value_type == datetime → ISO-8601 string anchored to event_ts
          5. Semantic pattern match on field name → domain-appropriate value
          6. value_type == number → round(random.uniform(1, 1000), 2)
          7. default → random 8-char alphanumeric string
        """
        vt = prop.value_type.lower()

        if prop.enum_values:
            return random.choice(prop.enum_values)

        if vt in ("boolean", "bool"):
            return random.choice([True, False])

        if vt == "list":
            if prop.enum_values:
                return random.choice(prop.enum_values)
            semantic = _semantic_value(prop.name, event_ts)
            return semantic if semantic is not None else _rand_string(8)

        if vt == "datetime":
            base = event_ts or _dt.datetime.now()
            delta = _dt.timedelta(
                days=random.randint(0, 365), seconds=random.randint(0, 86400)
            )
            return (base + delta).strftime("%Y-%m-%dT%H:%M:%S")

        # Semantic match covers both string and number fields
        semantic = _semantic_value(prop.name, event_ts)
        if semantic is not None:
            return semantic

        if vt in ("number", "int", "integer", "float", "double"):
            return round(random.uniform(1, 1000), 2)

        return _rand_string(8)


# ---------------------------------------------------------------------------
# CLI smoke test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    xlsx = (
        sys.argv[1]
        if len(sys.argv) > 1
        else "refrences/Annex 6 - Tracking Plan - Mini Program_V0.1.xlsx"
    )

    plan = TrackingPlan(xlsx)
    print(f"Events: {len(plan.list_events())}")
    print(f"Public props: {[p.name for p in plan.get_public_properties()]}")
    print(f"User attrs: {len(plan.get_user_attributes())}")

    schema = plan.get_event_schema("Registration_Result")
    if schema:
        print(f"Registration_Result props: {[p.name for p in schema.properties]}")
    else:
        print("Registration_Result not found")

    # Extra: show a few generated values
    print("\nSample generated values:")
    if schema:
        for prop in schema.properties[:5]:
            val = plan.generate_value(prop)
            print(f"  {prop.name} ({prop.value_type}): {val!r}")
