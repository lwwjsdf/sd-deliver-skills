#!/usr/bin/env python3
"""
generate_mock_data.py — 从埋点方案 Excel 生成模拟数据

用法：
    python3 tracking-setup-e2e/scripts/generate_mock_data.py
    python3 tracking-setup-e2e/scripts/generate_mock_data.py --count 100 --users 30

前置条件（在项目根目录的 .env 中配置）：
    SA_PROJECT          项目 ID
    TRACKING_PLAN_PATH  埋点方案 Excel 路径

依赖：
    pip install openpyxl python-dotenv
"""

import argparse
import base64
import json
import os
import random
import string
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from dotenv import load_dotenv
except ImportError:
    print("缺少依赖，请先运行: pip install openpyxl python-dotenv")
    sys.exit(1)

# .env 查找顺序：当前目录 → 父目录 → 祖父目录
for _p in [Path.cwd(), Path.cwd().parent, Path.cwd().parent.parent]:
    if (_p / ".env").exists():
        load_dotenv(_p / ".env")
        break

SA_PROJECT = os.getenv("SA_PROJECT", "default")
TRACKING_PLAN_PATH = os.getenv("TRACKING_PLAN_PATH", "")


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def parse_tracking_plan(xlsx_path: str) -> dict:
    """Parse a Sensors Data tracking plan Excel into a structured dict."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    plan = {"events": [], "users": []}

    # --- Events sheet ---
    if "Events" in wb.sheetnames:
        ws = wb["Events"]
        rows = list(ws.iter_rows(values_only=True))

        # Find header row (contains 'Event Variable Name')
        header_row_idx = None
        for i, row in enumerate(rows):
            if any(cell == "Event Variable Name" for cell in row if cell):
                header_row_idx = i
                break

        if header_row_idx is not None:
            headers = rows[header_row_idx]
            col = {v: i for i, v in enumerate(headers) if v}

            current_event = None
            for row in rows[header_row_idx + 1:]:
                event_name = row[col.get("Event Variable Name", 1)] if col.get("Event Variable Name") is not None else None
                attr_name = row[col.get("Event  Attribute Variable Name", 3)] if col.get("Event  Attribute Variable Name") is not None else None
                attr_display = row[col.get("Event  Attribute Variable Display Name", 4)] if col.get("Event  Attribute Variable Display Name") is not None else None
                data_type = row[col.get("Date Type", 5)] if col.get("Date Type") is not None else None
                sample = row[col.get("Sample Data", 11)] if col.get("Sample Data") is not None else None
                remark = row[col.get("Remark", 12)] if col.get("Remark") is not None else None

                if event_name:
                    current_event = {
                        "name": event_name,
                        "display_name": row[col.get("Event Display Name", 2)] if col.get("Event Display Name") is not None else event_name,
                        "properties": []
                    }
                    plan["events"].append(current_event)

                if current_event and attr_name:
                    options = _parse_options(remark) or _parse_options(sample)
                    current_event["properties"].append({
                        "name": str(attr_name).strip(),
                        "display_name": str(attr_display or attr_name).strip(),
                        "type": str(data_type or "String").strip(),
                        "sample": sample,
                        "options": options,
                    })

    # --- Users sheet ---
    if "Users" in wb.sheetnames:
        ws = wb["Users"]
        rows = list(ws.iter_rows(values_only=True))

        header_row_idx = None
        for i, row in enumerate(rows):
            if any(cell == "Attribute variable name" for cell in row if cell):
                header_row_idx = i
                break

        if header_row_idx is not None:
            headers = rows[header_row_idx]
            col = {v: i for i, v in enumerate(headers) if v}

            for row in rows[header_row_idx + 1:]:
                attr_name = row[col.get("Attribute variable name", 1)] if col.get("Attribute variable name") is not None else None
                if attr_name:
                    sample = row[col.get("Sample Data", 9)] if col.get("Sample Data") is not None else None
                    remark = row[col.get("Remark", 10)] if col.get("Remark") is not None else None
                    options = _parse_options(remark) or _parse_options(sample)
                    plan["users"].append({
                        "name": str(attr_name).strip(),
                        "display_name": str(row[col.get("Attribute display name", 2)] or attr_name).strip() if col.get("Attribute display name") is not None else str(attr_name).strip(),
                        "type": str(row[col.get("Data Type", 3)] or "String").strip() if col.get("Data Type") is not None else "String",
                        "sample": sample,
                        "options": options,
                    })

    return plan


# ---------------------------------------------------------------------------
# Mock value generation
# ---------------------------------------------------------------------------

def _parse_options(text: str) -> list[str]:
    """从 Remark/Sample 文本中提取值域列表（支持分号和换行分隔）。"""
    if not text:
        return []
    text = str(text).strip()
    # 换行分隔优先（Remark 列常见格式）
    if "\n" in text:
        opts = [s.strip() for s in text.splitlines() if s.strip()]
    else:
        opts = [s.strip() for s in text.split(";") if s.strip()]
    return opts if len(opts) > 1 else []


def _extract_prop(row: tuple, col: dict, name_key: str, display_key: str,
                  type_key: str, sample_key: str, remark_key: str) -> dict:
    """从一行 Excel 数据提取属性定义，包含值域信息。"""
    name = row[col.get(name_key, 1)] if col.get(name_key) is not None else None
    if not name:
        return {}
    display = row[col.get(display_key)] if col.get(display_key) is not None else None
    dtype = row[col.get(type_key)] if col.get(type_key) is not None else None
    sample = row[col.get(sample_key)] if col.get(sample_key) is not None else None
    remark = row[col.get(remark_key)] if col.get(remark_key) is not None else None

    # 值域：优先从 Remark 提取，其次从 Sample 提取（分号分隔的多值）
    options = _parse_options(remark) or _parse_options(sample)

    return {
        "name": str(name).strip(),
        "display_name": str(display or name).strip(),
        "type": str(dtype or "String").strip(),
        "sample": sample,
        "options": options,
    }


def _random_string(length=8):
    return "".join(random.choices(string.ascii_letters + string.digits, k=length))


def _random_datetime(days_back=90):
    delta = timedelta(days=random.randint(0, days_back), seconds=random.randint(0, 86400))
    return (datetime.now() - delta).strftime("%Y-%m-%dT%H:%M:%S.000Z")


def _random_timestamp_ms(days_back=90):
    delta = timedelta(days=random.randint(0, days_back), seconds=random.randint(0, 86400))
    return int((datetime.now() - delta).timestamp() * 1000)


def _timestamp_for_day(day_offset: int) -> int:
    """day_offset=0 是今天，day_offset=29 是 30 天前。当天随机时刻。"""
    base = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    base -= timedelta(days=day_offset)
    base += timedelta(seconds=random.randint(0, 86399))
    return int(base.timestamp() * 1000)


def generate_value(prop: dict) -> object:
    """根据属性定义生成模拟值，优先使用值域列表，其次 sample，最后按类型 fallback。"""
    dtype = (prop.get("type") or "String").strip().lower()
    sample = prop.get("sample")
    options = prop.get("options") or []

    # 有值域列表时直接随机选（最真实）
    if options:
        return random.choice(options)

    # sample 是单个具体值时直接用
    if sample is not None and not isinstance(sample, bool):
        sample_str = str(sample).strip()
        if sample_str and len(sample_str) < 80 and "\n" not in sample_str:
            return sample

    # 按类型 fallback
    if dtype in ("bool", "boolean", "checkbox"):
        return random.choice([True, False])

    if dtype in ("datetime", "date/time", "date"):
        return _random_datetime()

    if dtype in ("number", "int", "integer", "double", "currency"):
        return round(random.uniform(1, 10000), 2)

    if dtype in ("list", "picklist", "picklist (multi-select)", "set"):
        return f"Option_{random.choice('ABCD')}"

    # Default: String
    return _random_string(12)


def generate_user_id() -> str:
    return f"uid_{random.randint(1, 1000000):07d}"


def generate_identities(uid: str, idx: int) -> dict:
    return {
        "$identity_login_id": uid,
        "$identity_mobile": str(13600000000 + idx),
        "$identity_email": f"email_{idx:05d}@gmail.com",
    }


# ---------------------------------------------------------------------------
# Sensors Data event record builder
# ---------------------------------------------------------------------------

def build_track_record(event: dict, distinct_id: str, event_time_ms: int,
                       user_idx: int, project: str) -> dict:
    properties = {
        "$app_version": "1.0.0",
        "$lib": "python",
        "$lib_version": "1.0.0",
    }
    for prop in event.get("properties", []):
        properties[prop["name"]] = generate_value(prop)

    return {
        "distinct_id": distinct_id,
        "login_id": distinct_id,
        "type": "track",
        "event": event["name"],
        "time": event_time_ms,
        "time_free": True,
        "$is_login_id": True,
        "project": project,
        "identities": generate_identities(distinct_id, user_idx),
        "properties": properties,
    }


def build_profile_record(user_attrs: list, distinct_id: str,
                         user_idx: int, project: str) -> dict:
    properties = {}
    for attr in user_attrs:
        properties[attr["name"]] = generate_value(attr)
    return {
        "distinct_id": distinct_id,
        "login_id": distinct_id,
        "type": "profile_set",
        "time": int(time.time() * 1000),
        "time_free": True,
        "$is_login_id": True,
        "project": project,
        "identities": generate_identities(distinct_id, user_idx),
        "properties": properties,
    }


# ---------------------------------------------------------------------------
# Output formats
# ---------------------------------------------------------------------------

def to_batch_payload(records: list) -> str:
    """Encode records as base64 for Sensors Data batch HTTP API."""
    data = json.dumps(records, ensure_ascii=False, cls=_JSONEncoder)
    return base64.b64encode(data.encode("utf-8")).decode("utf-8")


class _JSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.strftime("%Y-%m-%dT%H:%M:%S.000Z")
        if hasattr(obj, "isoformat"):
            return obj.isoformat()
        if hasattr(obj, "__str__"):
            return str(obj)
        return super().default(obj)


def write_outputs(records: list, output_dir: str, prefix: str):
    os.makedirs(output_dir, exist_ok=True)

    # Raw JSON (one record per line, for readability and streaming import)
    jsonl_path = os.path.join(output_dir, f"{prefix}.jsonl")
    with open(jsonl_path, "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False, cls=_JSONEncoder) + "\n")

    # Batch payload (base64, ready to POST to /sa?token=xxx)
    batch_path = os.path.join(output_dir, f"{prefix}_batch.txt")
    with open(batch_path, "w", encoding="utf-8") as f:
        data = json.dumps(records, ensure_ascii=False, cls=_JSONEncoder)
        f.write(base64.b64encode(data.encode("utf-8")).decode("utf-8"))

    return jsonl_path, batch_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="从埋点方案 Excel 生成模拟数据")
    parser.add_argument("--count", type=int, default=None, help="每个事件生成的记录数（简单模式）")
    parser.add_argument("--users", type=int, default=None, help="模拟用户总数")
    parser.add_argument("--days", type=int, default=None, help="覆盖天数（时序模式）")
    parser.add_argument("--daily-users", type=int, default=None, help="每天活跃用户数（时序模式，需配合 --days）")
    parser.add_argument("--daily-count", type=int, default=None, help="每用户每天每事件记录数（时序模式，默认 100）")
    parser.add_argument("--output", default=None, help="输出目录（默认：脚本所在目录的 ../mock_data）")
    args = parser.parse_args()

    # 验证必要配置
    if not TRACKING_PLAN_PATH:
        print("错误：缺少必要配置，请在 .env 中设置 TRACKING_PLAN_PATH")
        sys.exit(1)

    excel_path = Path(TRACKING_PLAN_PATH)
    if not excel_path.exists():
        print(f"错误：找不到埋点方案文件：{excel_path}")
        sys.exit(1)

    # 输出目录
    output_dir = args.output or str(Path(__file__).parent.parent / "mock_data")

    plan = parse_tracking_plan(str(excel_path))
    events = plan["events"]
    user_attrs = plan["users"]

    # 用户池
    users_n = args.users if args.users is not None else 20
    users = [(generate_user_id(), i + 1) for i in range(users_n)]

    all_records = []

    # User profile records
    if user_attrs:
        for uid, idx in users:
            all_records.append(build_profile_record(user_attrs, uid, idx, SA_PROJECT))

    use_time_series = args.days is not None or args.daily_users is not None

    if use_time_series:
        # 时序模式：每天选 daily_users 个活跃用户，每人每事件生成 daily_count 条
        days = args.days if args.days is not None else 30
        daily_users_n = args.daily_users if args.daily_users is not None else min(20, users_n)
        daily_count = args.daily_count if args.daily_count is not None else 100

        print(f"=== 模拟数据生成（时序模式）===")
        print(f"文件: {excel_path.name}")
        print(f"项目: {SA_PROJECT}  用户池: {users_n}  每日活跃: {daily_users_n}  天数: {days}  每用户每事件: {daily_count}\n")
        print(f"发现 {len(events)} 个事件，{len(user_attrs)} 个用户属性")

        for day_offset in range(days):
            daily_active = random.sample(users, min(daily_users_n, len(users)))
            for event in events:
                for uid, idx in daily_active:
                    for _ in range(daily_count):
                        ts = _timestamp_for_day(day_offset)
                        all_records.append(build_track_record(event, uid, ts, idx, SA_PROJECT))
    else:
        # 简单模式：每事件生成 count 条，随机时间
        count = args.count if args.count is not None else 50

        print(f"=== 模拟数据生成 ===")
        print(f"文件: {excel_path.name}")
        print(f"项目: {SA_PROJECT}  每事件记录数: {count}  模拟用户数: {users_n}\n")
        print(f"发现 {len(events)} 个事件，{len(user_attrs)} 个用户属性")

        for event in events:
            for _ in range(count):
                uid, idx = random.choice(users)
                ts = _random_timestamp_ms()
                all_records.append(build_track_record(event, uid, ts, idx, SA_PROJECT))

    random.shuffle(all_records)

    jsonl_path, batch_path = write_outputs(all_records, output_dir, "mock_events")

    print(f"\n生成 {len(all_records)} 条记录")
    print(f"  JSONL:       {jsonl_path}")
    print(f"  Batch (b64): {batch_path}")
    print(f"\n导入命令（需要 SA_TOKEN）：")
    print(f"  curl -X POST \"$SA_HOST/sa?token=$SA_TOKEN\" \\")
    print(f"       -d 'data='$(cat {batch_path})")


if __name__ == "__main__":
    main()
