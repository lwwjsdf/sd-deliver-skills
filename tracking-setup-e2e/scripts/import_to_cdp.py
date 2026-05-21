#!/usr/bin/env python3
"""
import_to_cdp.py — 从埋点方案 Excel 自动导入元事件和用户属性到神策 CDP

用法：
    python3 tracking-setup-e2e/scripts/import_to_cdp.py

前置条件（在项目根目录的 .env 中配置）：
    SA_HOST             神策 CDP 地址
    SA_PROJECT          项目 ID
    TRACKING_PLAN_PATH  埋点方案 Excel 路径

依赖：
    pip install openpyxl python-dotenv
"""

import json
import os
import subprocess
import sys
import time
from pathlib import Path

try:
    import openpyxl
    from dotenv import load_dotenv
except ImportError:
    print("缺少依赖，请先运行: pip install openpyxl python-dotenv")
    sys.exit(1)

# .env 查找顺序：当前目录 → 父目录 → 祖父目录（兼容从任意位置运行）
for _p in [Path.cwd(), Path.cwd().parent, Path.cwd().parent.parent]:
    if (_p / ".env").exists():
        load_dotenv(_p / ".env")
        break

SA_HOST = os.getenv("SA_HOST", "").rstrip("/")
SA_PROJECT = os.getenv("SA_PROJECT", "")
TRACKING_PLAN_PATH = os.getenv("TRACKING_PLAN_PATH", "")
def _find_browse() -> Path:
    candidates = [
        Path.home() / ".claude/skills/gstack/browse/dist/browse",
        Path.home() / ".hermes/skills/gstack/browse/dist/browse",
        Path.home() / ".agents/skills/gstack/browse/dist/browse",
    ]
    # Also check if installed alongside this script (skill-local copy)
    candidates.append(Path(__file__).parent.parent.parent / "gstack/browse/dist/browse")
    for p in candidates:
        if p.exists():
            return p
    return candidates[0]  # return default path so error message is useful

BROWSE_BIN = _find_browse()


def validate_env(required: list[str]):
    missing = [k for k in required if not os.getenv(k)]
    if missing:
        print(f"错误：缺少必要配置，请在 .env 中设置：{', '.join(missing)}")
        sys.exit(1)


def _browse(cmd: list[str]) -> str:
    result = subprocess.run([str(BROWSE_BIN)] + cmd, capture_output=True, text=True)
    return result.stdout.strip()


def ensure_session():
    """通过 gstack/browse cookie picker 自动导入 Chrome 登录态"""
    if not BROWSE_BIN.exists():
        print("错误：未找到 gstack/browse")
        print("安装方式：npx skills add gstack/browse -g")
        print(f"查找路径：{BROWSE_BIN}")
        sys.exit(1)

    hostname = SA_HOST.replace("https://", "").replace("http://", "").split("/")[0]

    # 尝试带 --domain 参数直接导入（跳过 UI 选择器）
    result = subprocess.run(
        [str(BROWSE_BIN), "cookie-import-browser", "Chrome", "--domain", hostname],
        capture_output=True, text=True, timeout=15,
    )
    if result.returncode == 0:
        print(f"  ✓ 登录态导入成功（{hostname}）")
        _browse(["goto", f"{SA_HOST}/report/?project={SA_PROJECT}"])
        time.sleep(2)
        return

    # 回退：通过 cookie picker UI 自动操作
    proc = subprocess.Popen(
        [str(BROWSE_BIN), "cookie-import-browser", "Chrome"],
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
    )
    picker_url = None
    try:
        for line in proc.stdout:
            if "cookie-picker" in line:
                picker_url = line.strip().split()[-1]
                break
    except Exception:
        pass
    if not picker_url:
        print("  ⚠ 无法获取 cookie picker URL，跳过登录态导入")
        return
    time.sleep(0.5)
    _browse(["goto", picker_url])
    time.sleep(0.5)
    _browse(["fill", "[placeholder='Search domains...']", hostname])
    time.sleep(0.5)
    text = _browse(["text"])
    if hostname in text:
        print(f"  ✓ 登录态导入成功（{hostname}）")
    _browse(["goto", f"{SA_HOST}/report/?project={SA_PROJECT}"])
    time.sleep(2)


def fetch(path: str, method: str = "GET", body: dict = None) -> dict:
    """通过 browse 发起 API 请求，复用已登录的 session"""
    url = f"{SA_HOST}{path}"
    if body:
        js = (
            f"fetch('{url}',{{method:'{method}',credentials:'include',"
            f"headers:{{'Content-Type':'application/json'}},"
            f"body:JSON.stringify({json.dumps(body, ensure_ascii=False)})"
            f"}}).then(r=>r.json()).then(d=>JSON.stringify(d))"
        )
    else:
        js = f"fetch('{url}',{{credentials:'include'}}).then(r=>r.json()).then(d=>JSON.stringify(d))"
    raw = _browse(["js", js])
    try:
        return json.loads(raw)
    except Exception:
        return {"_raw": raw}

# ── 数据类型映射 ──────────────────────────────────────────────
TYPE_MAP = {
    "String": "STRING",
    "List": "LIST",
    "Datetime": "DATETIME",
    "Bool": "BOOL",
    "Number": "NUMBER",
}

# 系统保留字段名（无法创建，自动跳过）
RESERVED_FIELD_NAMES = {"Id", "PersonEmail"}


# ── Excel 解析 ────────────────────────────────────────────────
def parse_events(wb: openpyxl.Workbook) -> list[dict]:
    """从 Events sheet 解析元事件列表。

    搜索 'Event Variable Name' 表头行作为起点，跳过前面的 Identity 配置区。
    """
    if "Events" not in wb.sheetnames:
        return []
    ws = wb["Events"]
    rows = list(ws.iter_rows(values_only=True))

    # 找到列标题行（含 'Event Variable Name'）
    header_row_idx = None
    for i, row in enumerate(rows):
        if any(str(c).strip() == "Event Variable Name" for c in row if c):
            header_row_idx = i
            break

    if header_row_idx is None:
        # 回退：从第2行开始，跳过非 int 序号行
        data_rows = rows[1:]
    else:
        data_rows = rows[header_row_idx + 1:]

    events = []
    for row in data_rows:
        if not row[0] or not isinstance(row[0], int):
            continue
        original_name = row[1]
        display_name = row[2] if len(row) > 2 else original_name
        if not original_name:
            continue
        events.append({
            "original_name": str(original_name).strip(),
            "display_name": str(display_name or original_name).strip(),
        })
    return events


def parse_event_fields(wb: openpyxl.Workbook, event_name: str) -> list[dict]:
    """从 Details（Event） sheet 解析事件属性列表。

    支持两种结构：
    1. 按事件分区：每个事件有独立标题行，cell0 == event_name 标记起点
    2. 共享属性库：所有事件共用一套属性（无事件名标题行），直接读全部
    """
    sheet_name = next(
        (s for s in wb.sheetnames if s.startswith("Details") and "Event" in s),
        None,
    )
    if not sheet_name:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # 找列标题行（含 'Attribute variable name' 或 'Event  Attribute Variable Name'）
    header_keywords = {"attribute variable name", "event  attribute variable name"}
    header_idx = None
    for i, row in enumerate(rows):
        if any(str(c).strip().lower() in header_keywords for c in row if c):
            header_idx = i
            break

    if header_idx is None:
        return []

    # 检查是否有事件分区（event_name 出现在 col0）
    has_event_sections = any(
        str(row[0]).strip() == event_name
        for row in rows[header_idx + 1:]
        if row[0] and not isinstance(row[0], int)
    )

    fields = []
    if has_event_sections:
        # 按事件分区：找到 event_name 标题行后读属性
        in_event = False
        for row in rows[header_idx + 1:]:
            cell0 = str(row[0]).strip() if row[0] else ""
            if cell0 == event_name:
                in_event = True
                continue
            if not in_event:
                continue
            serial, name, display, dtype = row[0], row[1], row[2], row[3]
            if not isinstance(serial, int):
                if serial:
                    break  # 下一个事件的标题行
                continue
            if not name:
                continue
            name = str(name).strip()
            if name in RESERVED_FIELD_NAMES:
                print(f"    跳过保留字段名: {name}")
                continue
            fields.append({
                "display_name": str(display or name).strip(),
                "name": name,
                "data_type": TYPE_MAP.get(str(dtype), "STRING"),
                "field_type": "BASIC",
                "unit_remark": "",
                "remark": "",
            })
    else:
        # 共享属性库：所有事件共用，直接读全部属性
        for row in rows[header_idx + 1:]:
            serial, name, display, dtype = row[0], row[1], row[2], row[3]
            if not isinstance(serial, int) or not name:
                continue
            name = str(name).strip()
            if name in RESERVED_FIELD_NAMES:
                print(f"    跳过保留字段名: {name}")
                continue
            fields.append({
                "display_name": str(display or name).strip(),
                "name": name,
                "data_type": TYPE_MAP.get(str(dtype), "STRING"),
                "field_type": "BASIC",
                "unit_remark": "",
                "remark": "",
            })

    return fields


def parse_user_attrs(wb: openpyxl.Workbook) -> list[dict]:
    """从 Users sheet 解析用户属性列表（第2列=属性名，第3列=显示名，第4列=类型）"""
    if "Users" not in wb.sheetnames:
        return []
    ws = wb["Users"]
    fields = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        serial, name, display, dtype = row[0], row[1], row[2], row[3]
        if not name or not isinstance(serial, int):
            continue
        name = str(name).strip()
        if name in RESERVED_FIELD_NAMES:
            print(f"  跳过保留字段名: {name}")
            continue
        fields.append({
            "display_name": str(display or name).strip(),
            "name": name,
            "data_type": TYPE_MAP.get(str(dtype), "STRING"),
            "field_type": "BASIC",
            "unit_remark": "",
            "remark": "",
        })
    return fields


# ── API 操作 ──────────────────────────────────────────────────
def create_meta_event(event: dict, fields: list[dict]) -> bool:
    """创建元事件并写入自定义属性"""
    name = event["original_name"]

    # Step 1: 创建事件（带 appKey 基础字段）
    resp = fetch(
        f"/api/v2/horizon/v1/web/event_schema/create?event_type=META&project={SA_PROJECT}",
        method="POST",
        body={
            "event_type": "META",
            "entity_name": "user",
            "meta": {
                "physical_name": "events",
                "entity_name": "user",
                "event_type": "META",
                "original_name": name,
                "display_name": event["display_name"],
                "need_track_platforms": [{"name": "MINI_APP"}],
                "fields": [{
                    "display_name": "appKey", "name": "appKey",
                    "data_type": "STRING", "field_type": "BASIC",
                    "visible": False, "unit_remark": "", "remark": "",
                }],
            },
        },
    )
    if resp.get("code") != "SUCCESS":
        print(f"  ✗ 创建失败: {str(resp.get('message') or resp.get('error', ''))[:100]}")
        return False

    if not fields:
        print(f"  ✓ 创建成功（无自定义属性）")
        return True

    # Step 2: 获取现有字段（保留虚拟字段，排除 preset/builtin）
    detail = fetch(
        f"/api/v2/horizon/v1/web/event_schema/detail"
        f"?event_type=META&project={SA_PROJECT}&name=events.{name}&entity_name=user"
    )
    existing = detail.get("data", {}).get("meta", {}).get("fields", [])
    keep = [f for f in existing if not f.get("is_preset") and not f.get("is_builtin")]

    # Step 3: update 补充自定义属性
    resp2 = fetch(
        f"/api/v2/horizon/v1/web/event_schema/update"
        f"?event_type=META&project={SA_PROJECT}&name=events.{name}&entity_name=user",
        method="POST",
        body={
            "event_type": "META",
            "entity_name": "user",
            "meta": {
                "physical_name": "events",
                "entity_name": "user",
                "event_type": "META",
                "name": f"events.{name}",
                "original_name": name,
                "display_name": event["display_name"],
                "need_track_platforms": [{"name": "MINI_APP"}],
                "fields": keep + fields,
            },
        },
    )
    if resp2.get("code") == "SUCCESS":
        print(f"  ✓ 创建成功，写入 {len(fields)} 个属性")
        return True
    else:
        print(f"  ✗ 属性写入失败: {str(resp2.get('message', ''))[:100]}")
        return False


def insert_user_attrs(fields: list[dict]) -> dict:
    """逐个插入用户属性（避免单个失败影响全部）"""
    ok, skipped, failed = [], [], []
    for f in fields:
        resp = fetch(
            f"/api/v2/horizon/v1/web/schema/field/batch_insert?project={SA_PROJECT}",
            method="POST",
            body={"fields": [f], "schema_class": "USER", "physical_name": "users"},
        )
        code = resp.get("code", "")
        msg = str(resp.get("message") or resp.get("error", ""))
        if code == "SUCCESS":
            ok.append(f["name"])
        elif "ALREADY_EXISTS" in msg or "already exist" in msg.lower():
            skipped.append(f["name"])
        else:
            failed.append({"name": f["name"], "error": msg[:80]})
        time.sleep(0.1)

    print(f"  新增: {len(ok)}  已存在: {len(skipped)}  失败: {len(failed)}")
    for f in failed:
        print(f"    ✗ {f['name']}: {f['error']}")
    return {"ok": ok, "skipped": skipped, "failed": failed}


# ── 主流程 ────────────────────────────────────────────────────
def main():
    validate_env(["SA_HOST", "SA_PROJECT", "TRACKING_PLAN_PATH"])

    excel_path = Path(TRACKING_PLAN_PATH)
    if not excel_path.exists():
        print(f"错误：找不到埋点方案文件：{excel_path}")
        sys.exit(1)

    print(f"=== 神策 CDP 元数据导入 ===")
    print(f"文件: {excel_path.name}")
    print(f"目标: {SA_HOST}  项目: {SA_PROJECT}\n")

    ensure_session()

    wb = openpyxl.load_workbook(excel_path)
    results = {"events": [], "user_attrs": {}}

    # ── 元事件 ──
    print("── 元事件导入 ──")
    events = parse_events(wb)
    if not events:
        print("  未找到 Events sheet 或无数据，跳过")
    else:
        print(f"  发现 {len(events)} 个事件")
        for event in events:
            print(f"\n  [{event['original_name']}]")
            fields = parse_event_fields(wb, event["original_name"])
            success = create_meta_event(event, fields)
            results["events"].append({
                "name": event["original_name"],
                "success": success,
                "fields": len(fields),
            })

    # ── 用户属性 ──
    print("\n── 用户属性导入 ──")
    user_fields = parse_user_attrs(wb)
    if not user_fields:
        print("  未找到 Users sheet 或无数据，跳过")
    else:
        print(f"  发现 {len(user_fields)} 个属性")
        results["user_attrs"] = insert_user_attrs(user_fields)

    # ── 汇总 ──
    print("\n=== 导入完成 ===")
    if results["events"]:
        ok_n = sum(1 for e in results["events"] if e["success"])
        print(f"元事件: {ok_n}/{len(results['events'])} 成功")
    if results["user_attrs"]:
        ua = results["user_attrs"]
        print(f"用户属性: 新增 {len(ua.get('ok', []))}，"
              f"已存在 {len(ua.get('skipped', []))}，"
              f"失败 {len(ua.get('failed', []))}")


if __name__ == "__main__":
    main()
