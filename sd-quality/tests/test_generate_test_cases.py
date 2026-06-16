"""Tests for generate_test_cases.py."""
import openpyxl
import pytest

from generate_test_cases import generate_sit_test_cases, generate_uat_test_cases


def _make_minimal_tracking_plan(path: str):
    wb = openpyxl.Workbook()

    # Custom Event sheet: col 1=event, 3=attr, 5=type, 7=trigger
    ws = wb.active
    ws.title = "Custom Event"
    ws.append(["", "event", "event display", "attr", "attr display", "type", "example", "trigger", "", "", ""])
    ws.append(["", "Login", "登录", "isSuccess", "是否成功", "boolean", "true", "MP", "", "", ""])
    ws.append(["", "", "", "userType", "用户类型", "string", "vip", "MP", "", "", ""])

    # Public Property sheet: header has "attribute english variable name"; col 0=attr, 3=type, 4=example
    ws2 = wb.create_sheet("Public  Property")
    ws2.append(["attribute english variable name", "", "", "type", "example", "", "", "", "", "", ""])
    ws2.append(["platformType", "", "", "string", "MP", "", "", "", "", "", ""])

    # User Attribute sheet: col 0=attr, 2=type, 3=example
    ws3 = wb.create_sheet("User Attribute")
    ws3.append(["attr", "", "type", "example", "", ""])
    ws3.append(["registerTime", "", "datetime", "2025-01-01", "", ""])

    wb.save(path)


def test_generate_sit_test_cases(tmp_path):
    tp = tmp_path / "tp.xlsx"
    out = tmp_path / "sit.xlsx"
    _make_minimal_tracking_plan(str(tp))

    path, count = generate_sit_test_cases(str(tp), str(out))

    assert path == str(out)
    assert count > 0
    assert out.exists()

    wb = openpyxl.load_workbook(out, data_only=True)
    assert wb.sheetnames == ["SIT Test Cases"]
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "Test Case ID" in headers
    assert "Priority" in headers

    descriptions = [ws.cell(row=r, column=headers.index("Test Case Description") + 1).value for r in range(2, ws.max_row + 1)]
    assert any("Login" in str(v) for v in descriptions)


def test_generate_uat_test_cases(tmp_path):
    tp = tmp_path / "tp.xlsx"
    out = tmp_path / "uat.xlsx"
    _make_minimal_tracking_plan(str(tp))

    path, sheets = generate_uat_test_cases(str(tp), str(out))

    assert path == str(out)
    assert set(sheets) == {"Indicators", "ID-Mapping", "Permissions", "Paths"}
    assert out.exists()

    wb = openpyxl.load_workbook(out, data_only=True)
    for name in sheets:
        assert name in wb.sheetnames

    ws = wb["Indicators"]
    headers = [cell.value for cell in ws[1]]
    assert "Indicator Name" in headers


def test_generate_sit_test_cases_empty_tracking_plan(tmp_path):
    tp = tmp_path / "tp.xlsx"
    out = tmp_path / "sit.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Custom Event"
    wb.create_sheet("Public  Property")
    wb.create_sheet("User Attribute")
    wb.save(tp)

    path, count = generate_sit_test_cases(str(tp), str(out))
    assert count == 0
    assert out.exists()
