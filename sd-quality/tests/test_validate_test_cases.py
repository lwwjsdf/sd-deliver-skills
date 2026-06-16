"""Tests for validate_test_cases.py."""
import openpyxl
import pytest

from validate_test_cases import (
    SIT_REQUIRED_HEADERS,
    UAT_SHEETS,
    validate_sit,
    validate_uat,
)


def _make_valid_sit(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SIT Test Cases"
    ws.append(SIT_REQUIRED_HEADERS)
    ws.append(["1.0.0", "Verify Login", "TC-001", "Ready", "Trigger", "Event in CDP", "P0", "", "", "", ""])
    wb.save(path)


def _make_valid_uat(path):
    wb = openpyxl.Workbook()
    for sheet_name, headers in UAT_SHEETS.items():
        if wb.active.title == "Sheet":
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        ws.append(["x"] * len(headers))
    wb.save(path)


def test_validate_sit_valid(tmp_path):
    p = tmp_path / "sit.xlsx"
    _make_valid_sit(p)
    assert validate_sit(str(p)) == []


def test_validate_sit_missing_sheet(tmp_path):
    p = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Wrong Sheet"
    wb.save(p)
    errors = validate_sit(str(p))
    assert any("缺少 sheet" in e for e in errors)


def test_validate_sit_missing_header(tmp_path):
    p = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SIT Test Cases"
    ws.append(["Test Case ID", "Priority"])
    wb.save(p)
    errors = validate_sit(str(p))
    assert any("SIT 列缺失" in e for e in errors)


def test_validate_sit_no_data(tmp_path):
    p = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SIT Test Cases"
    ws.append(SIT_REQUIRED_HEADERS)
    wb.save(p)
    errors = validate_sit(str(p))
    assert any("没有数据行" in e for e in errors)


def test_validate_uat_valid(tmp_path):
    p = tmp_path / "uat.xlsx"
    _make_valid_uat(p)
    assert validate_uat(str(p)) == []


def test_validate_uat_missing_sheet(tmp_path):
    p = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Indicators"
    wb.active.append(UAT_SHEETS["Indicators"])
    wb.active.append(["x"] * len(UAT_SHEETS["Indicators"]))
    wb.save(p)
    errors = validate_uat(str(p))
    assert any("缺少 sheet" in e for e in errors)


def test_validate_uat_wrong_headers(tmp_path):
    p = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    for sheet_name in UAT_SHEETS:
        if wb.active.title == "Sheet":
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        ws.append(["Wrong", "Header"])
    wb.save(p)
    errors = validate_uat(str(p))
    assert any("列缺失" in e for e in errors)
