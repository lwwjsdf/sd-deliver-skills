#!/usr/bin/env python3
"""
validate_test_cases.py — Validate SIT or UAT Test Case Excel structure.

Usage:
    ./venv/bin/python <skill-repo>/sd-quality/scripts/validate_test_cases.py \
      --type sit \
      --input ./references/sit-test-case.xlsx

    ./venv/bin/python <skill-repo>/sd-quality/scripts/validate_test_cases.py \
      --type uat \
      --input ./references/uat-test-case.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Dict, List

try:
    import openpyxl
except ImportError as e:
    print("Missing dependency: openpyxl")
    print("Run: ./venv/bin/pip install -r <skill-repo>/requirements.txt")
    sys.exit(1)


SIT_SHEET_NAME = "SIT Test Cases"
SIT_REQUIRED_HEADERS = [
    "Release Number",
    "Test Case Description",
    "Test Case ID",
    "Precondition",
    "Step Number and Description",
    "Expected Result",
    "Priority",
    "Status",
    "Tester",
    "Test Date",
    "Bug/Remark",
]

UAT_SHEETS: Dict[str, List[str]] = {
    "Indicators": [
        "Scenario",
        "Case No.",
        "Indicator Name",
        "Indicator Definition",
        "Related Events",
        "Formula",
        "Expected Result",
        "Tester",
        "Test Date",
        "Status",
        "Bug/Remark",
    ],
    "ID-Mapping": [
        "Scenario",
        "Test Case No.",
        "Test Approach",
        "Expected Result",
        "Tester",
        "Test Date",
        "Status",
        "Screencap File Name",
        "Bug/Remark",
    ],
    "Permissions": [
        "BU",
        "Data Role Permission",
        "Function Role Permission",
        "Dashboard Permission",
        "Case",
        "Expected Result",
        "Status",
    ],
    "Paths": [
        "Scenario",
        "Case No.",
        "Path Steps",
        "Expected Events",
        "Expected Result",
        "Status",
    ],
}


def _read_headers(ws) -> List[str]:
    if ws.max_row < 1:
        return []
    return [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]


def _clean_headers(headers: List[str]) -> List[str]:
    return [h for h in headers if h]


def validate_sit(path: str) -> List[str]:
    errors: List[str] = []
    wb = openpyxl.load_workbook(path, data_only=True)

    if SIT_SHEET_NAME not in wb.sheetnames:
        errors.append(f"缺少 sheet '{SIT_SHEET_NAME}'，实际 sheets: {wb.sheetnames}")
        return errors

    ws = wb[SIT_SHEET_NAME]
    headers = _clean_headers(_read_headers(ws))

    if headers != SIT_REQUIRED_HEADERS:
        missing = [h for h in SIT_REQUIRED_HEADERS if h not in headers]
        extra = [h for h in headers if h not in SIT_REQUIRED_HEADERS]
        if missing:
            errors.append(f"SIT 列缺失: {missing}")
        if extra:
            errors.append(f"SIT 列多余: {extra}")
        if headers != SIT_REQUIRED_HEADERS and not missing and not extra:
            errors.append(f"SIT 列顺序/内容不匹配。期望: {SIT_REQUIRED_HEADERS}，实际: {headers}")

    if ws.max_row < 2:
        errors.append("SIT sheet 没有数据行")

    return errors


def validate_uat(path: str) -> List[str]:
    errors: List[str] = []
    wb = openpyxl.load_workbook(path, data_only=True)

    for sheet_name, required_headers in UAT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            errors.append(f"缺少 sheet '{sheet_name}'，实际 sheets: {wb.sheetnames}")
            continue

        ws = wb[sheet_name]
        headers = _clean_headers(_read_headers(ws))

        if headers != required_headers:
            missing = [h for h in required_headers if h not in headers]
            extra = [h for h in headers if h not in required_headers]
            if missing:
                errors.append(f"UAT sheet '{sheet_name}' 列缺失: {missing}")
            if extra:
                errors.append(f"UAT sheet '{sheet_name}' 列多余: {extra}")
            if headers != required_headers and not missing and not extra:
                errors.append(
                    f"UAT sheet '{sheet_name}' 列顺序/内容不匹配。"
                    f"期望: {required_headers}，实际: {headers}"
                )

        if ws.max_row < 2:
            errors.append(f"UAT sheet '{sheet_name}' 没有数据行")

    return errors


def main():
    parser = argparse.ArgumentParser(description="Validate SIT/UAT Test Case Excel")
    parser.add_argument("--type", choices=["sit", "uat"], required=True)
    parser.add_argument("--input", required=True, help="Path to test case Excel")
    args = parser.parse_args()

    path = args.input
    if not Path(path).exists():
        print(f"❌ File not found: {path}")
        sys.exit(1)

    if args.type == "sit":
        errors = validate_sit(path)
    else:
        errors = validate_uat(path)

    if errors:
        print(f"❌ {args.type.upper()} Test Case validation failed:")
        for err in errors:
            print(f"  - {err}")
        sys.exit(1)

    print(f"✅ {args.type.upper()} Test Case structure is valid.")


if __name__ == "__main__":
    main()
