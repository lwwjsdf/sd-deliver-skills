#!/usr/bin/env python3
"""
generate_test_cases.py — Generate SIT or UAT Test Case Excel from a Tracking Plan.

Usage:
    python3 generate_test_cases.py \
      --type sit \
      --tracking-plan ./references/tracking-plan.xlsx \
      --output ./references/sit-test-case.xlsx

    python3 generate_test_cases.py \
      --type uat \
      --tracking-plan ./references/tracking-plan.xlsx \
      --output ./references/uat-test-case.xlsx

Dependencies:
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import re
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Optional


def _ensure_dependencies():
    missing = []
    try:
        import openpyxl
    except ImportError:
        missing.append("openpyxl")
    if missing:
        print(f"Missing dependencies: {', '.join(missing)}")
        print("Installing...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install"] + missing)
            print("Dependencies installed successfully.")
        except subprocess.CalledProcessError as e:
            print(f"Failed to install dependencies: {e}")
            print(f"Please run manually: pip install {' '.join(missing)}")
            sys.exit(1)


_ensure_dependencies()

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).parent))
from tracking_plan import TrackingPlan


# ---------------------------------------------------------------------------
# SIT generation
# ---------------------------------------------------------------------------


def _sit_rows(tracking_plan: TrackingPlan) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    counter = 1

    public_props = tracking_plan.get_public_properties()
    public_names = [p.name for p in public_props]

    for event_name in sorted(tracking_plan.list_events()):
        schema = tracking_plan.get_event_schema(event_name)
        if not schema:
            continue

        # One row to verify the event itself
        rows.append({
            "Release Number": "1.0.0",
            "Test Case Description": f"Verify {event_name} event capture",
            "Test Case ID": f"TC-{counter:03d}",
            "Precondition": "SDK integrated and test environment ready",
            "Step Number and Description": f"1. Trigger {event_name}\n2. Verify event appears in CDP",
            "Expected Result": f"Event '{event_name}' is received with correct timestamp and identities",
            "Priority": "P0",
            "Status": "",
        })
        counter += 1

        # One row per event property
        for prop in schema.properties:
            if prop.name in public_names:
                continue
            priority = "P0" if prop.required else "P1"
            rows.append({
                "Release Number": "1.0.0",
                "Test Case Description": f"Verify {event_name}.{prop.name} ({prop.value_type})",
                "Test Case ID": f"TC-{counter:03d}",
                "Precondition": f"Event {event_name} exists",
                "Step Number and Description": f"1. Trigger {event_name}\n2. Check property '{prop.name}'",
                "Expected Result": f"Property '{prop.name}' is present, type is {prop.value_type}",
                "Priority": priority,
                "Status": "",
            })
            counter += 1

    # Public property coverage
    for prop in public_props:
        rows.append({
            "Release Number": "1.0.0",
            "Test Case Description": f"Verify public property {prop.name} on all events",
            "Test Case ID": f"TC-{counter:03d}",
            "Precondition": "Any event exists",
            "Step Number and Description": "1. Trigger multiple events\n2. Check public property presence",
            "Expected Result": f"Property '{prop.name}' is present on every event",
            "Priority": "P0",
            "Status": "",
        })
        counter += 1

    return rows


# ---------------------------------------------------------------------------
# UAT generation
# ---------------------------------------------------------------------------


def _uat_indicator_rows(tracking_plan: TrackingPlan) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    counter = 1

    event_names = sorted(tracking_plan.list_events())
    has_pageview = any("$pageview" in e for e in event_names)
    has_mp_launch = any("$MPLaunch" in e for e in event_names)
    has_purchase = any(
        re.search(r"order|payment|purchase", e, re.I) for e in event_names
    )

    if has_pageview:
        rows.append({
            "Scenario": "Traffic",
            "Case No.": counter,
            "Indicator Name": "Website Unique Visitors",
            "Indicator Definition": "Unique visitors within period",
            "Related Events": "$pageview",
            "Formula": "COUNT DISTINCT distinct_id WHERE platformType=Web",
            "Expected Result": "Match expected count",
            "Tester": "",
            "Test Date": "",
            "Status": "",
        })
        counter += 1

    if has_mp_launch:
        rows.append({
            "Scenario": "Traffic",
            "Case No.": counter,
            "Indicator Name": "Mini Program DAU",
            "Indicator Definition": "Daily active users",
            "Related Events": "$MPLaunch",
            "Formula": "COUNT DISTINCT distinct_id WHERE event=$MPLaunch per day",
            "Expected Result": "Match expected count",
            "Tester": "",
            "Test Date": "",
            "Status": "",
        })
        counter += 1

    if has_purchase:
        purchase_event = next(
            (e for e in event_names if re.search(r"order.*payment|purchase", e, re.I)),
            "Purchase",
        )
        rows.append({
            "Scenario": "Conversion",
            "Case No.": counter,
            "Indicator Name": "Purchase Conversion Rate",
            "Indicator Definition": "Orders / Visitors",
            "Related Events": f"$pageview, {purchase_event}",
            "Formula": f"COUNT({purchase_event}) / COUNT(DISTINCT $pageview.distinct_id)",
            "Expected Result": "Match expected rate",
            "Tester": "",
            "Test Date": "",
            "Status": "",
        })
        counter += 1

    # Fallback: ensure at least one indicator for any non-preset event
    if not rows and event_names:
        target = next((e for e in event_names if not e.startswith("$")), event_names[0])
        rows.append({
            "Scenario": "Event Count",
            "Case No.": counter,
            "Indicator Name": f"{target} Count",
            "Indicator Definition": f"Total count of {target}",
            "Related Events": target,
            "Formula": f"COUNT {target}",
            "Expected Result": "Match expected count",
            "Tester": "",
            "Test Date": "",
            "Status": "",
        })

    return rows


def _uat_id_mapping_rows() -> List[Dict[str, str]]:
    return [
        {
            "Scenario": "Cross-platform",
            "Test Case No.": 1,
            "Test Approach": "Same user logs in on MP and Web; query by email in Users Look-Up",
            "Expected Result": "One user record contains both MP and Web events",
            "Tester": "",
            "Test Date": "",
            "Status": "",
            "Screencap File Name": "",
            "Bug/Remark": "",
        },
        {
            "Scenario": "MP only",
            "Test Case No.": 2,
            "Test Approach": "Query by mobile in Users Look-Up",
            "Expected Result": "One record; $identity_mobile and $identity_mp_unionid have values",
            "Tester": "",
            "Test Date": "",
            "Status": "",
            "Screencap File Name": "",
            "Bug/Remark": "",
        },
    ]


def _uat_permission_rows() -> List[Dict[str, str]]:
    return [
        {
            "BU": "M+",
            "Data Role Permission": "M+ Data Only",
            "Function Role Permission": "Dashboard Viewer",
            "Dashboard Permission": "M+ All Dashboard",
            "Case": "Data isolation",
            "Expected Result": "Only M+ data visible",
            "Status": "",
        }
    ]


def _uat_path_rows(tracking_plan: TrackingPlan) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    event_names = sorted(tracking_plan.list_events())
    for event_name in event_names:
        if event_name.startswith("$"):
            continue
        rows.append({
            "Scenario": f"Path: {event_name}",
            "Case No.": len(rows) + 1,
            "Path Steps": f"Trigger {event_name}",
            "Expected Events": event_name,
            "Expected Result": f"Event '{event_name}' captured with all required properties",
            "Status": "",
        })
    return rows


# ---------------------------------------------------------------------------
# Excel output helpers
# ---------------------------------------------------------------------------


def _style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="00A870", end_color="00A870", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_sheet(ws, rows: List[Dict[str, str]]):
    if not rows:
        ws.append(["No data"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        _style_header(cell)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 60)


def generate_sit_test_cases(tracking_plan_path: str, output_path: str):
    plan = TrackingPlan(tracking_plan_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "SIT Test Cases"
    rows = _sit_rows(plan)
    _write_sheet(ws, rows)
    wb.save(output_path)
    return output_path, len(rows)


def generate_uat_test_cases(tracking_plan_path: str, output_path: str):
    plan = TrackingPlan(tracking_plan_path)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Indicators"
    _write_sheet(ws1, _uat_indicator_rows(plan))

    ws2 = wb.create_sheet("ID-Mapping")
    _write_sheet(ws2, _uat_id_mapping_rows())

    ws3 = wb.create_sheet("Permissions")
    _write_sheet(ws3, _uat_permission_rows())

    ws4 = wb.create_sheet("Paths")
    _write_sheet(ws4, _uat_path_rows(plan))

    wb.save(output_path)
    return output_path, wb.sheetnames


def main():
    parser = argparse.ArgumentParser(description="Generate SIT/UAT Test Case Excel from Tracking Plan")
    parser.add_argument("--type", choices=["sit", "uat"], required=True, help="Test case type")
    parser.add_argument("--tracking-plan", required=True, help="Path to Tracking Plan Excel")
    parser.add_argument("--output", required=True, help="Output Excel path")
    args = parser.parse_args()

    if args.type == "sit":
        path, count = generate_sit_test_cases(args.tracking_plan, args.output)
        print(f"SIT Test Case generated: {path}")
        print(f"Total test cases: {count}")
    else:
        path, sheets = generate_uat_test_cases(args.tracking_plan, args.output)
        print(f"UAT Test Case generated: {path}")
        print(f"Sheets: {', '.join(sheets)}")


if __name__ == "__main__":
    main()
