"""Tests for gen_excel.py."""
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from gen_excel import (
    apply_arm_double,
    fill_deploy_info,
    fill_eval_page,
    fill_server_info,
    parse_node_groups,
    safe_set,
    set_cell,
    write_node_row,
)


def test_parse_node_groups():
    groups = parse_node_groups(["meta:1.1.1.1,1.1.1.2", "data:2.2.2.2"])
    assert groups[0]["type"] == "meta"
    assert groups[0]["ips"] == ["1.1.1.1", "1.1.1.2"]
    assert groups[1]["type"] == "data"


def test_parse_node_groups_default_hybrid():
    groups = parse_node_groups(["3.3.3.3"])
    assert groups[0]["type"] == "hybrid"


def test_apply_arm_double():
    result = apply_arm_double(1000, 500, 400, 250)
    assert result == (2000, 1000, 800, 500)


def test_set_cell_writes_value():
    wb = openpyxl.Workbook()
    ws = wb.active
    set_cell(ws, "A1", "hello")
    assert ws["A1"].value == "hello"


def test_set_cell_writes_to_merged_cell_master():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:B2")
    set_cell(ws, "B2", "merged")
    assert ws["A1"].value == "merged"


def test_safe_set_skips_merged_slave():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:B2")
    # B2 is a MergedCell, should not raise
    safe_set(ws, 2, 2, "value")


def test_write_node_row_meta():
    wb = openpyxl.Workbook()
    ws = wb.active
    write_node_row(ws, 5, "meta", "10.0.0.1", 1000, 500, 400, 250)
    assert ws.cell(row=5, column=1).value == "meta"
    assert ws.cell(row=5, column=2).value == "10.0.0.1"
    assert ws.cell(row=5, column=10).value == "400G"


def test_write_node_row_hybrid():
    wb = openpyxl.Workbook()
    ws = wb.active
    write_node_row(ws, 5, "hybrid", "10.0.0.1", 1000, 500, 400, 250, seq_disk_count=3)
    assert ws.cell(row=5, column=1).value == "hybrid"
    assert ws.cell(row=5, column=11).value == "/dev/vdc"
    assert ws.cell(row=5, column=12).value == "1000G"
    assert ws.cell(row=6, column=11).value == "/dev/vdd"


def test_fill_eval_page_single():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "0.评估单元"
    # Mark expected cells
    for addr in ["C3", "C4", "A10", "B10", "C10"]:
        ws[addr] = ""

    params = {
        "cloud_vendor": "阿里云",
        "dau_range": "<100万",
        "daily_import": 0.5,
        "retention_days": 365,
        "history_events": 0,
    }
    fill_eval_page(wb, "single", params)
    assert ws["C3"].value == "阿里云"
    assert ws["B10"].value == 365


def test_fill_deploy_info():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3.服务器信息"
    for addr in ["P7", "P8", "P9", "P10", "P13"]:
        ws[addr] = ""

    params = {"cloud_region": "华北1", "ntp": "ntp.aliyun.com", "timezone": "Asia/Shanghai", "login_method": "ssh"}
    fill_deploy_info(wb, "standard", params)
    assert ws["P8"].value == "华北1"
    assert ws["P13"].value == "ssh"


def test_fill_server_info_single():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3.服务器信息"
    fill_server_info(wb, "single", [{"type": "hybrid", "ips": ["10.0.0.1"]}],
                     1000, 500, 400, 250, 3)
    assert ws.cell(row=18, column=1).value == "hybrid"
    assert ws.cell(row=18, column=2).value == "10.0.0.1"


def test_fill_server_info_standard():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3.服务器信息"
    node_groups = [
        {"type": "meta", "ips": ["10.0.0.1", "10.0.0.2", "10.0.0.3"]},
        {"type": "data", "ips": ["10.0.0.4"]},
    ]
    fill_server_info(wb, "standard", node_groups, 1000, 500, 400, 250, 3)
    assert ws.cell(row=18, column=1).value == "meta"
    assert ws.cell(row=21, column=1).value == "data"


def test_generate_excel_uses_template(monkeypatch, tmp_path):
    """generate_excel should copy template, fill values, and save."""
    from gen_excel import generate_excel

    fake_template = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "0.评估页"
    wb.create_sheet("3.服务器信息")
    wb.save(fake_template)

    monkeypatch.setattr("gen_excel.TEMPLATES", {"mini": fake_template})

    output = tmp_path / "out.xlsx"
    generate_excel(
        mode="mini",
        output_path=str(output),
        cloud_vendor="阿里云",
        dau_range="<100万",
        daily_import=0.5,
        retention_days=365,
        history_events=0,
        node_groups=[{"type": "hybrid", "ips": ["10.0.0.1", "10.0.0.2", "10.0.0.3"]}],
        seq_disk_size=1000,
        rand_disk_size=500,
        meta_disk_size=400,
        staging_disk_size=250,
        seq_disk_count=3,
    )
    assert output.exists()

    wb_out = openpyxl.load_workbook(output, data_only=True)
    assert "0.评估页" in wb_out.sheetnames
    assert "3.服务器信息" in wb_out.sheetnames
