"""Tests for design_performance_test.py."""
from pathlib import Path

import openpyxl
import pytest

from design_performance_test import generate_word_report, generate_excel_report


def test_generate_excel_report(tmp_path):
    out = tmp_path / "perf.xlsx"
    path = generate_excel_report(
        output_path=str(out),
        dau=1_000_000,
        daily_events=5_000_000,
        retention_days=365,
        cloud="AWS",
        region="ap-southeast-1",
        include_cdp=True,
        include_ma=True,
    )
    assert path == str(out)
    assert out.exists()

    wb = openpyxl.load_workbook(out, data_only=True)
    assert "Scope" in wb.sheetnames
    assert "Environment" in wb.sheetnames
    assert "Assumptions" in wb.sheetnames
    assert "Schedule" in wb.sheetnames

    ws = wb["Scope"]
    headers = [cell.value for cell in ws[1]]
    assert "Scenario" in headers


def test_generate_word_report(tmp_path):
    out = tmp_path / "perf.docx"
    path = generate_word_report(
        output_path=str(out),
        dau=1_000_000,
        daily_events=5_000_000,
        retention_days=365,
        cloud="AWS",
        region="ap-southeast-1",
        include_cdp=True,
        include_ma=True,
    )
    assert path == str(out)
    assert out.exists()

    from docx import Document
    doc = Document(str(out))
    texts = [p.text for p in doc.paragraphs]
    assert any("Performance Test Plan" in t for t in texts)
    assert any("Scope of Testing" in t for t in texts)


def test_generate_markdown_report(tmp_path):
    from design_performance_test import generate_markdown_report

    out = tmp_path / "perf.md"
    path = generate_markdown_report(
        output_path=str(out),
        dau=1_000_000,
        daily_events=5_000_000,
        retention_days=365,
        cloud="AWS",
        region="ap-southeast-1",
        include_cdp=True,
        include_ma=True,
    )
    assert path == str(out)
    assert out.exists()
    content = out.read_text(encoding="utf-8")
    assert "# Performance Test Plan" in content
    assert "PT-001" in content
    assert "AWS" in content


def test_cli_requires_at_least_one_output(tmp_path, monkeypatch):
    import design_performance_test

    import sys
    old_argv = sys.argv
    try:
        sys.argv = ["design_performance_test.py", "--dau", "1000000", "--daily-events", "5000000"]
        with pytest.raises(SystemExit):
            design_performance_test.main()
    finally:
        sys.argv = old_argv
