#!/usr/bin/env python3
"""
神策服务器资源评估 Excel 生成器
基于官方模板复制填值，保留所有格式、公式、合并单元格。

用法：
  python3 gen_excel.py \
    --mode single|mini|standard \
    --cloud-vendor 阿里云 \
    --dau-range "<100万" \
    --daily-import 0.5 \
    --retention-days 365 \
    --history-events 0 \
    --data-nodes 3 \
    --output /path/to/output.xlsx \
    [--nodes "meta:10.0.0.1,10.0.0.2,10.0.0.3" \
     --nodes "data:10.0.0.4,10.0.0.5,10.0.0.6"] \
    [--seq-disk-size 1000 \
     --rand-disk-size 500 \
     --meta-disk-size 400] \
    [--cloud-region 华北1 \
     --ntp ntp.aliyun.com \
     --timezone Asia/Shanghai \
     --login-method "ssh直连"]
"""

import argparse
import shutil
import os
import sys
from pathlib import Path
from copy import copy

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("错误：需要安装 openpyxl。运行：pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# 模板路径（相对于本脚本所在目录）
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).parent
REPO_ROOT = SCRIPT_DIR.parent

TEMPLATES = {
    "single":    REPO_ROOT / "refrences" / "公有云SA 单机模版_V3.7.2_20251125.xlsx",
    "mini":      REPO_ROOT / "refrences" / "公有云SA mini 集群模版_V3.7.2_20251125.xlsx",
    "standard":  REPO_ROOT / "refrences" / "公有云SA 标准集群模版_V3.7.2_20251125.xlsx",
    "sfn-3node": REPO_ROOT / "refrences" / "SFN 3 节点集群v2.4.4.xlsx",
    "sfn-3+3":   REPO_ROOT / "refrences" / "SFN 3+3 集群 v2.4.4.xlsx",
    "sfn-3+n+3": REPO_ROOT / "refrences" / "SFN 3+N+3 集群 v2.4.4.xlsx",
}

# ---------------------------------------------------------------------------
# 评估页填值映射
# ---------------------------------------------------------------------------

# 单机：sheet "0.评估单元"
SINGLE_EVAL_CELLS = {
    "cloud_vendor":    "C3",   # 云厂商
    "dau_range":       "C4",   # 平均日活范围
    "daily_import":    "A10",  # 日导入数据条数（亿）
    "retention_days":  "B10",  # 存储天数
    "history_events":  "C10",  # 历史已有事件量（亿）
}

# Mini 集群：sheet "0.评估页"
MINI_EVAL_CELLS = {
    "cloud_vendor":    "D3",
    "dau_range":       "D4",
    "daily_import":    "B10",
    "retention_days":  "C10",
    "history_events":  "D10",
}

# 标准集群：sheet "0.评估页"
STANDARD_EVAL_CELLS = {
    "cloud_vendor":    "D3",
    "dau_range":       "D4",
    "daily_import":    "B11",
    "retention_days":  "C11",
    "history_events":  "D11",
}

# SFN 3节点 / 3+3 / 3+N+3：sheet "0.评估单元"
SFN_EVAL_CELLS = {
    "cloud_vendor":    "C3",   # 云厂商（下拉）
    "dau_range":       "C4",   # 平均日活范围（下拉）
}

# ---------------------------------------------------------------------------
# 服务器信息页 — 列定义
# ---------------------------------------------------------------------------

# 列字母 → 字段含义
SERVER_INFO_COLS = {
    "A": "node_type",
    "B": "ip",
    "C": "hostname",
    "D": "user",
    "E": "passwd",
    "F": "port",
    "G": "system_disk_dev",
    "H": "system_disk_size",
    "I": "meta_disk_dev",
    "J": "meta_disk_size",
    "K": "seq_disk_dev",
    "L": "seq_disk_size",
    "M": "rand_disk_dev",
    "N": "rand_disk_size",
    "O": "staging_disk_dev",
    "P": "staging_disk_size",
    "Q": "remark",
}

# 部署信息区（右侧，两种集群模版结构相同）
DEPLOY_INFO_CELLS = {
    "cloud_vendor":   "P7",
    "cloud_region":   "P8",
    "ntp":            "P9",
    "timezone":       "P10",
    "login_method":   "P13",
}

# 单机模版部署信息区列不同（L/M列）
SINGLE_DEPLOY_INFO_CELLS = {
    "cloud_vendor":   "M7",
    "cloud_region":   "M8",
    "ntp":            "M9",
    "timezone":       "M10",
    "login_method":   "M13",
}

# ---------------------------------------------------------------------------
# 节点行起始位置（CDP 模板）
# ---------------------------------------------------------------------------

# mini 集群：hybrid 节点，每节点占 3 行（主行 + 2 续行）
MINI_NODE_START_ROW = 18
MINI_NODE_ROW_SPAN = 3   # 每节点占行数

# 标准集群：meta 节点每节点 1 行，data 节点每节点 3 行
STANDARD_META_START_ROW = 18
STANDARD_DATA_START_ROW = 21
STANDARD_DATA_ROW_SPAN = 3

# 单机：hybrid 节点，1 行
SINGLE_NODE_ROW = 18

# ---------------------------------------------------------------------------
# SFN 模板服务器信息页结构常量
# ---------------------------------------------------------------------------

# SFN 3节点：混合节点×3，每节点 4 行（标题行 + 3 行数据）
SFN_3NODE_NODE_START_ROWS = [10, 14, 18]   # node01/node02/node03 的标题行
SFN_3NODE_DATA_START_ROWS = [11, 15, 19]   # 对应数据起始行
SFN_3NODE_SEQ_COUNT = 3

# SFN 3+3：元数据节点 3 个（每节点 2 行：标题+数据），SF混合节点 3 个（每节点 4 行：标题+3数据）
SFN_3PLUS3_META_TITLE_ROWS = [11, 13, 15]
SFN_3PLUS3_META_DATA_ROWS  = [12, 14, 16]
SFN_3PLUS3_SF_TITLE_ROWS   = [27, 31, 35]
SFN_3PLUS3_SF_DATA_ROWS    = [28, 32, 36]
SFN_3PLUS3_SEQ_COUNT = 3

# SFN 3+N+3：元数据节点 3 个（每节点 2 行），数据节点 N 个（每节点 4 行），SF在线节点 3 个（每节点 4 行）
SFN_3N3_META_TITLE_ROWS = [11, 13, 15]
SFN_3N3_META_DATA_ROWS  = [12, 14, 16]
SFN_3N3_DATA_TITLE_START_ROW = 27   # data01 标题行
SFN_3N3_DATA_DATA_START_ROW  = 28   # data01 数据起始行
SFN_3N3_SF_TITLE_START_ROW   = 59   # SF在线节点 data01 标题行（基于 3 个数据节点模板）
SFN_3N3_SF_DATA_START_ROW    = 60   # SF在线节点数据起始行
SFN_3N3_SEQ_COUNT = 3

# ---------------------------------------------------------------------------
# 辅助函数
# ---------------------------------------------------------------------------

def set_cell(ws, cell_addr: str, value):
    """设置单元格值，跳过合并单元格的从属格。"""
    cell = ws[cell_addr]
    # 如果是合并单元格的从属格，找到主格
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # 只写主格（左上角）
            min_col = merged_range.min_col
            min_row = merged_range.min_row
            ws.cell(row=min_row, column=min_col).value = value
            return
    cell.value = value


def fill_eval_page(wb, mode: str, params: dict):
    """填写评估页的橙色输入单元格。"""
    if mode == "single":
        sheet_name = "0.评估单元"
        cell_map = SINGLE_EVAL_CELLS
    elif mode in ("sfn-3node", "sfn-3+3", "sfn-3+n+3"):
        sheet_name = "0.评估单元"
        cell_map = SFN_EVAL_CELLS
    else:
        sheet_name = "0.评估页"
        cell_map = MINI_EVAL_CELLS if mode == "mini" else STANDARD_EVAL_CELLS

    ws = wb[sheet_name]

    if params.get("cloud_vendor"):
        set_cell(ws, cell_map["cloud_vendor"], params["cloud_vendor"])
    if params.get("dau_range"):
        set_cell(ws, cell_map["dau_range"], params["dau_range"])
    # SFN 模板评估页只有云厂商和日活两个输入项，其余字段跳过
    if mode in ("sfn-3node", "sfn-3+3", "sfn-3+n+3"):
        return
    if params.get("daily_import") is not None:
        set_cell(ws, cell_map["daily_import"], float(params["daily_import"]))
    if params.get("retention_days") is not None:
        set_cell(ws, cell_map["retention_days"], int(params["retention_days"]))
    if params.get("history_events") is not None:
        set_cell(ws, cell_map["history_events"], float(params["history_events"]))


def fill_deploy_info(wb, mode: str, params: dict):
    """填写服务器信息页右侧的部署信息区。"""
    ws = wb["3.服务器信息"]
    cell_map = SINGLE_DEPLOY_INFO_CELLS if mode == "single" else DEPLOY_INFO_CELLS

    for key, addr in cell_map.items():
        if params.get(key):
            set_cell(ws, addr, params[key])


def safe_set(ws, row: int, col: int, value):
    """写入单元格。如果目标格是合并格的从属格，直接跳过（不写入）。"""
    cell = ws.cell(row=row, column=col)
    if cell.__class__.__name__ == "MergedCell":
        return  # 从属格不可写，跳过
    cell.value = value


def write_node_row(ws, row: int, node_type: str, ip: str,
                   seq_disk_size: int, rand_disk_size: int,
                   meta_disk_size: int, staging_disk_size: int,
                   seq_disk_count: int = 3,
                   user: str = "root", passwd: str = "", port: int = 22):
    """
    写入一个节点的主行及续行（顺序盘续行）。
    seq_disk_count: 顺序盘数量（mini/data 节点通常 3 或 6 块）
    返回实际占用的行数。
    """
    # 主行
    safe_set(ws, row, 1, node_type)
    safe_set(ws, row, 2, ip)
    safe_set(ws, row, 4, user)
    safe_set(ws, row, 5, passwd)
    safe_set(ws, row, 6, port)

    # 系统盘（所有节点都有）
    safe_set(ws, row, 7, "/dev/vda1")
    safe_set(ws, row, 8, "50G")

    if node_type == "meta":
        # 元数据节点：只有系统盘 + 元数据盘
        safe_set(ws, row, 9, "/dev/vdb")
        safe_set(ws, row, 10, f"{meta_disk_size}G")
        return 1  # 占 1 行

    # data / hybrid 节点
    if node_type in ("data", "hybrid"):
        # 元数据盘（hybrid 有，data 无）
        if node_type == "hybrid":
            safe_set(ws, row, 9, "/dev/vdb")
            safe_set(ws, row, 10, f"{meta_disk_size}G")
            seq_start_dev = "c"
            rand_start_idx = ord("c") + seq_disk_count
        else:
            seq_start_dev = "b"
            rand_start_idx = ord("b") + seq_disk_count

        # 顺序盘（主行写第1块，续行写剩余）
        dev_letters = [chr(ord(seq_start_dev) + i) for i in range(seq_disk_count)]
        safe_set(ws, row, 11, f"/dev/vd{dev_letters[0]}")
        safe_set(ws, row, 12, f"{seq_disk_size}G")

        for i, letter in enumerate(dev_letters[1:], start=1):
            safe_set(ws, row + i, 11, f"/dev/vd{letter}")
            safe_set(ws, row + i, 12, f"{seq_disk_size}G")

        # 随机数据盘（第1块写在主行）
        rand_dev = chr(rand_start_idx)
        safe_set(ws, row, 13, f"/dev/vd{rand_dev}")
        safe_set(ws, row, 14, f"{rand_disk_size}G")

        # 暂存盘
        staging_dev = chr(rand_start_idx + 1)
        safe_set(ws, row, 15, f"/dev/vd{staging_dev}")
        safe_set(ws, row, 16, f"{staging_disk_size}G")

        return seq_disk_count  # 占行数 = 顺序盘数量（主行 + seq_disk_count-1 续行）

    return 1


def _write_sfn_node_block(ws, title_row: int, data_start_row: int,
                          node_name: str, ip: str,
                          seq_disk_size: int, rand_disk_size: int,
                          meta_disk_size: int,
                          seq_disk_count: int = 3,
                          has_meta: bool = True,
                          has_seq: bool = True,
                          has_rand: bool = True,
                          user: str = "root", passwd: str = "", port: int = 22):
    """
    写入 SFN 模板中一个节点块（标题行 + 数据行）。
    SFN 模板结构：
      标题行: A=node名（如 node01/data01）
      数据行1: A=IP, B=user, C=passwd, D=system_dev, E=system_size,
               F=meta_dev, G=meta_size,
               H=seq1_dev, I=seq1_size, J=rand1_dev, K=rand1_size, L=remark
      数据行2+: H=seqN_dev, I=seqN_size, J=randN_dev, K=randN_size
    """
    # 标题行
    safe_set(ws, title_row, 1, node_name)

    # 数据行1
    row = data_start_row
    safe_set(ws, row, 1, ip)
    safe_set(ws, row, 2, user)
    safe_set(ws, row, 3, passwd)
    safe_set(ws, row, 4, "/dev/vda1")
    safe_set(ws, row, 5, "40G")

    if has_meta:
        safe_set(ws, row, 6, "/dev/vda2")
        safe_set(ws, row, 7, f"{meta_disk_size}G")
        seq_start_dev = "c"
        rand_start_dev_base = ord("c") + seq_disk_count
    else:
        seq_start_dev = "b"
        rand_start_dev_base = ord("b") + seq_disk_count

    if has_seq:
        dev_letters = [chr(ord(seq_start_dev) + i) for i in range(seq_disk_count)]
        safe_set(ws, row, 8, f"/dev/vd{dev_letters[0]}")
        safe_set(ws, row, 9, f"{seq_disk_size}G")
        for i, letter in enumerate(dev_letters[1:], start=1):
            safe_set(ws, row + i, 8, f"/dev/vd{letter}")
            safe_set(ws, row + i, 9, f"{seq_disk_size}G")

    if has_rand:
        rand_dev = chr(rand_start_dev_base)
        safe_set(ws, row, 10, f"/dev/vd{rand_dev}")
        safe_set(ws, row, 11, f"{rand_disk_size}G")
        # SFN 3节点/3+3 有 3 块随机盘，分 3 行写
        for i in range(1, seq_disk_count):
            rand_dev = chr(rand_start_dev_base + i)
            safe_set(ws, row + i, 10, f"/dev/vd{rand_dev}")
            safe_set(ws, row + i, 11, f"{rand_disk_size}G")

    safe_set(ws, row, 12, "盘符按实际情况填写")


def _fill_sfn_server_info(ws, mode: str, node_groups: list,
                          seq_disk_size: int, rand_disk_size: int,
                          meta_disk_size: int, seq_disk_count: int):
    """
    填写 SFN 模板服务器信息页。
    node_groups: [{"type": "meta/hybrid/data/sfonline", "ips": [...]}]
    """
    meta_ips = []
    data_ips = []      # 3+3 的 SF混合节点 或 3+N+3 的数据节点
    sfonline_ips = []  # 3+N+3 的 SF在线节点

    for g in node_groups:
        t = g["type"]
        if t == "meta":
            meta_ips = g["ips"]
        elif t in ("data", "hybrid"):
            data_ips = g["ips"]
        elif t == "sfonline":
            sfonline_ips = g["ips"]

    if mode == "sfn-3node":
        # 3节点：混合节点×3，每节点 4 行（标题+3数据行）
        ips = data_ips if data_ips else [""] * 3
        for i, ip in enumerate(ips[:3]):
            _write_sfn_node_block(
                ws, SFN_3NODE_NODE_START_ROWS[i], SFN_3NODE_DATA_START_ROWS[i],
                f"node{i+1:02d}", ip,
                seq_disk_size, rand_disk_size, meta_disk_size,
                seq_disk_count=seq_disk_count,
                has_meta=True, has_seq=True, has_rand=True,
            )
        return

    if mode == "sfn-3+3":
        # 元数据节点×3
        for i, ip in enumerate(meta_ips[:3]):
            _write_sfn_node_block(
                ws, SFN_3PLUS3_META_TITLE_ROWS[i], SFN_3PLUS3_META_DATA_ROWS[i],
                f"meta{i+1:02d}", ip,
                seq_disk_size, rand_disk_size, meta_disk_size,
                seq_disk_count=0,
                has_meta=True, has_seq=False, has_rand=False,
            )
        # SF 混合节点×3
        for i, ip in enumerate(data_ips[:3]):
            _write_sfn_node_block(
                ws, SFN_3PLUS3_SF_TITLE_ROWS[i], SFN_3PLUS3_SF_DATA_ROWS[i],
                f"data{i+1:02d}", ip,
                seq_disk_size, rand_disk_size, meta_disk_size,
                seq_disk_count=seq_disk_count,
                has_meta=False, has_seq=True, has_rand=True,
            )
        return

    if mode == "sfn-3+n+3":
        # 元数据节点×3
        for i, ip in enumerate(meta_ips[:3]):
            _write_sfn_node_block(
                ws, SFN_3N3_META_TITLE_ROWS[i], SFN_3N3_META_DATA_ROWS[i],
                f"meta{i+1:02d}", ip,
                seq_disk_size, rand_disk_size, meta_disk_size,
                seq_disk_count=0,
                has_meta=True, has_seq=False, has_rand=False,
            )

        # 数据节点：模板预留 3 个（data01-data03），每节点 4 行
        # 若实际节点数 >3，需插入行；若 <3，只填前几个
        template_data_nodes = 3
        actual_data_nodes = len(data_ips)
        data_title_start = SFN_3N3_DATA_TITLE_START_ROW
        data_data_start = SFN_3N3_DATA_DATA_START_ROW
        rows_per_data_node = 4  # 标题 + 3 数据行

        if actual_data_nodes > template_data_nodes:
            # 在分隔行前插入额外行
            separator_row = data_data_start + template_data_nodes * rows_per_data_node  # 43
            extra = (actual_data_nodes - template_data_nodes) * rows_per_data_node
            ws.insert_rows(separator_row, extra)
            to_remove = [
                mr for mr in list(ws.merged_cells.ranges)
                if separator_row <= mr.min_row <= separator_row + extra - 1
            ]
            for mr in to_remove:
                ws.merged_cells.remove(mr)

        for i, ip in enumerate(data_ips):
            title_row = data_title_start + i * rows_per_data_node
            data_row = data_data_start + i * rows_per_data_node
            _write_sfn_node_block(
                ws, title_row, data_row,
                f"data{i+1:02d}", ip,
                seq_disk_size, rand_disk_size, meta_disk_size,
                seq_disk_count=seq_disk_count,
                has_meta=False, has_seq=True, has_rand=True,
            )

        # SF 在线节点：模板中在数据节点之后
        # 基于模板 3 个数据节点时，SF在线节点标题从行 59 开始
        sf_title_start = data_data_start + actual_data_nodes * rows_per_data_node + 16  # 示例区占 16 行
        # 更精确：从模板已知位置计算。模板中 data03 数据结束在行 42，示例结束在行 57
        # SF在线 node01 标题在行 59，数据在行 60
        # 由于可能插入了数据节点行，需要重新计算
        # 简化：从最后一个数据节点的数据行 + 16（标题"SF 在线节点"+表头+示例区）
        # 实际上模板结构是固定的：数据节点区 → "示例"标题 → 3个示例 → "SF 在线节点"标题 → 表头 → 节点
        # 重新定位 "SF 在线节点" 标题行
        sf_section_row = None
        for r in range(data_data_start + actual_data_nodes * rows_per_data_node, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "SF 在线节点":
                sf_section_row = r
                break

        if sf_section_row:
            # SF 在线节点表头在 sf_section_row + 1，数据起始在 sf_section_row + 2
            sf_data_start = sf_section_row + 2
            for i, ip in enumerate(sfonline_ips[:3]):
                title_row = sf_data_start + i * rows_per_data_node
                data_row = title_row + 1
                _write_sfn_node_block(
                    ws, title_row, data_row,
                    f"data{i+1:02d}", ip,
                    seq_disk_size, rand_disk_size, meta_disk_size,
                    seq_disk_count=seq_disk_count,
                    has_meta=False, has_seq=False, has_rand=True,
                )
        return


def fill_server_info(wb, mode: str, node_groups: list,
                     seq_disk_size: int, rand_disk_size: int,
                     meta_disk_size: int, staging_disk_size: int,
                     seq_disk_count: int):
    """
    填写服务器信息页的节点行。
    node_groups: [{"type": "meta/data/hybrid/sfonline", "ips": [...]}]
    """
    if not node_groups:
        return

    ws = wb["3.服务器信息"]

    if mode in ("sfn-3node", "sfn-3+3", "sfn-3+n+3"):
        _fill_sfn_server_info(ws, mode, node_groups,
                              seq_disk_size, rand_disk_size, meta_disk_size, seq_disk_count)
        return

    if mode == "single":
        ips = node_groups[0]["ips"] if node_groups else []
        ip = ips[0] if ips else ""
        write_node_row(ws, SINGLE_NODE_ROW, "hybrid", ip,
                       seq_disk_size, rand_disk_size, meta_disk_size, staging_disk_size,
                       seq_disk_count=1)
        return

    if mode == "mini":
        # Mini：3 个 hybrid 节点，模板每节点预留 3 行
        # 若 seq_disk_count > 3，需要插入额外行
        ips = node_groups[0]["ips"] if node_groups else []
        rows_per_node = max(seq_disk_count, 3)  # 至少 3 行（模板预留）
        template_rows_per_node = 3

        current_row = MINI_NODE_START_ROW
        for i, ip in enumerate(ips[:3]):
            # 如果需要比模板多的行，先插入
            extra = rows_per_node - template_rows_per_node
            if extra > 0:
                ws.insert_rows(current_row + template_rows_per_node, extra)
            write_node_row(ws, current_row, "hybrid", ip,
                           seq_disk_size, rand_disk_size, meta_disk_size, staging_disk_size,
                           seq_disk_count=seq_disk_count)
            current_row += rows_per_node
        return

    if mode == "standard":
        meta_ips = []
        data_ips = []
        for g in node_groups:
            if g["type"] == "meta":
                meta_ips = g["ips"]
            elif g["type"] == "data":
                data_ips = g["ips"]

        # meta 节点（每节点 1 行，直接写）
        for i, ip in enumerate(meta_ips[:3]):
            row = STANDARD_META_START_ROW + i
            write_node_row(ws, row, "meta", ip,
                           seq_disk_size, rand_disk_size, meta_disk_size, staging_disk_size)

        # data 节点：模板预留 9 行（3节点×3行），计算实际需要的行数
        # 在合并格分隔行（行30）之前插入不足的行，并解除新行的合并格状态
        data_start = STANDARD_META_START_ROW + len(meta_ips[:3])
        template_data_rows = 9          # 模板预留：3节点×3行
        needed_data_rows = len(data_ips) * seq_disk_count
        extra_rows = max(0, needed_data_rows - template_data_rows)

        separator_row = data_start + template_data_rows  # 行30
        if extra_rows > 0:
            ws.insert_rows(separator_row, extra_rows)
            # 解除插入行上可能继承的合并格状态
            to_remove = [
                mr for mr in list(ws.merged_cells.ranges)
                if separator_row <= mr.min_row <= separator_row + extra_rows - 1
            ]
            for mr in to_remove:
                ws.merged_cells.remove(mr)

        # 顺序写所有 data 节点
        current_row = data_start
        for ip in data_ips:
            write_node_row(ws, current_row, "data", ip,
                           seq_disk_size, rand_disk_size, meta_disk_size, staging_disk_size,
                           seq_disk_count=seq_disk_count)
            current_row += seq_disk_count


# ---------------------------------------------------------------------------
# 主函数
# ---------------------------------------------------------------------------

def parse_node_groups(node_args: list) -> list:
    """
    解析 --nodes 参数，格式：'meta:10.0.0.1,10.0.0.2' 或 'data:10.0.0.3,10.0.0.4'
    返回：[{"type": "meta", "ips": [...]}, ...]
    """
    groups = []
    for arg in (node_args or []):
        if ":" in arg:
            node_type, ips_str = arg.split(":", 1)
            ips = [ip.strip() for ip in ips_str.split(",") if ip.strip()]
        else:
            # 没有类型前缀，默认 hybrid
            node_type = "hybrid"
            ips = [ip.strip() for ip in arg.split(",") if ip.strip()]
        groups.append({"type": node_type.strip(), "ips": ips})
    return groups


def apply_arm_double(seq_disk_size: int, rand_disk_size: int,
                     meta_disk_size: int, staging_disk_size: int) -> tuple:
    """ARM 架构：所有磁盘容量 double。"""
    return seq_disk_size * 2, rand_disk_size * 2, meta_disk_size * 2, staging_disk_size * 2


def generate_excel(
    mode: str,
    output_path: str,
    cloud_vendor: str,
    dau_range: str,
    daily_import: float,
    retention_days: int,
    history_events: float,
    node_groups: list,
    seq_disk_size: int,
    rand_disk_size: int,
    meta_disk_size: int,
    staging_disk_size: int,
    seq_disk_count: int,
    cloud_region: str = "",
    ntp: str = "ntp.aliyun.com",
    timezone: str = "Asia/Shanghai (东八区)",
    login_method: str = "",
    arch: str = "x86",
):
    template_path = TEMPLATES.get(mode)
    if not template_path or not template_path.exists():
        print(f"错误：找不到模板文件 {template_path}", file=sys.stderr)
        sys.exit(1)

    # ARM 架构：磁盘容量 double
    if arch.lower() == "arm":
        seq_disk_size, rand_disk_size, meta_disk_size, staging_disk_size = \
            apply_arm_double(seq_disk_size, rand_disk_size, meta_disk_size, staging_disk_size)

    # 复制模板
    shutil.copy2(template_path, output_path)

    # 用 openpyxl 打开副本（keep_vba=False，keep_links=False 避免外部引用警告）
    wb = openpyxl.load_workbook(output_path, keep_vba=False)

    params = {
        "cloud_vendor":   cloud_vendor,
        "dau_range":      dau_range,
        "daily_import":   daily_import,
        "retention_days": retention_days,
        "history_events": history_events,
        "cloud_region":   cloud_region,
        "ntp":            ntp,
        "timezone":       timezone,
        "login_method":   login_method,
    }

    fill_eval_page(wb, mode, params)
    fill_deploy_info(wb, mode, params)
    fill_server_info(wb, mode, node_groups,
                     seq_disk_size, rand_disk_size, meta_disk_size, staging_disk_size,
                     seq_disk_count)

    wb.save(output_path)
    print(f"✓ 已生成：{output_path}")


def main():
    parser = argparse.ArgumentParser(description="神策服务器资源评估 Excel 生成器")
    parser.add_argument("--mode",             required=True,
                        choices=["single", "mini", "standard",
                                 "sfn-3node", "sfn-3+3", "sfn-3+n+3"],
                        help="部署模式：single/mini/standard（CDP）或 sfn-3node/sfn-3+3/sfn-3+n+3（SFN）")
    parser.add_argument("--cloud-vendor",     required=True,  help="云厂商，如 阿里云、腾讯云、AWS")
    parser.add_argument("--dau-range",        required=True,  help="日活范围，如 '<100万'、'<800万'")
    parser.add_argument("--daily-import",     type=float, default=0,
                        help="日导入数据条数（亿），如 0.5")
    parser.add_argument("--retention-days",   type=int, default=365, help="存储天数（默认 365）")
    parser.add_argument("--history-events",   type=float, default=0,
                        help="历史已有事件量（亿），如 0")
    parser.add_argument("--nodes",            action="append", default=[],
                        help="节点 IP，格式：'meta:10.0.0.1,10.0.0.2' 或 'data:10.0.0.3'，可多次指定")
    parser.add_argument("--seq-disk-size",    type=int, default=1000,
                        help="顺序数据盘单块容量 GB（默认 1000）")
    parser.add_argument("--seq-disk-count",   type=int, default=3,
                        help="每节点顺序数据盘数量（默认 3，标准集群通常 6）")
    parser.add_argument("--rand-disk-size",   type=int, default=500,
                        help="随机数据盘单块容量 GB（默认 500）")
    parser.add_argument("--meta-disk-size",   type=int, default=400,
                        help="元数据盘容量 GB（默认 400）")
    parser.add_argument("--staging-disk-size",type=int, default=250,
                        help="暂存盘容量 GB（默认 250）")
    parser.add_argument("--cloud-region",     default="",     help="云服务器所在地域，如 华北1")
    parser.add_argument("--ntp",              default="ntp.aliyun.com", help="NTP 时钟源地址")
    parser.add_argument("--timezone",         default="Asia/Shanghai (东八区)", help="时区")
    parser.add_argument("--login-method",     default="",     help="远程登录方式，如 ssh直连")
    parser.add_argument("--arch",             default="x86",  help="CPU 架构：x86 或 arm（默认 x86）")
    parser.add_argument("--output",           required=True,  help="输出文件路径，如 /tmp/客户名_SA配置.xlsx")

    args = parser.parse_args()
    node_groups = parse_node_groups(args.nodes)

    generate_excel(
        mode=args.mode,
        output_path=args.output,
        cloud_vendor=args.cloud_vendor,
        dau_range=args.dau_range,
        daily_import=args.daily_import,
        retention_days=args.retention_days,
        history_events=args.history_events,
        node_groups=node_groups,
        seq_disk_size=args.seq_disk_size,
        seq_disk_count=args.seq_disk_count,
        rand_disk_size=args.rand_disk_size,
        meta_disk_size=args.meta_disk_size,
        staging_disk_size=args.staging_disk_size,
        cloud_region=args.cloud_region,
        ntp=args.ntp,
        timezone=args.timezone,
        login_method=args.login_method,
        arch=args.arch,
    )


if __name__ == "__main__":
    main()
