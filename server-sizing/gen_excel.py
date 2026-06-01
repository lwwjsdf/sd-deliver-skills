#!/usr/bin/env python3
"""
神策服务器资源评估 Excel 生成器
基于官方模板复制填值，保留所有格式、公式、合并单元格。

用法：
  python3 gen_excel.py \
    --mode single|mini|standard \
    --cloud-vendor 阿里云 \
    --dau-range "<100万" \
    --daily-import 0.5 \
    --retention-days 365 \
    --history-events 0 \
    --data-nodes 3 \
    --output /path/to/output.xlsx \
    [--nodes "meta:10.0.0.1,10.0.0.2,10.0.0.3" \
     --nodes "data:10.0.0.4,10.0.0.5,10.0.0.6"] \
    [--seq-disk-size 1000 \
     --rand-disk-size 500 \
     --meta-disk-size 400] \
    [--cloud-region 华北1 \
     --ntp ntp.aliyun.com \
     --timezone Asia/Shanghai \
     --login-method "ssh直连"]
"""

import argparse
import shutil
import os
import sys
from pathlib import Path
from copy import copy

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("错误：需要安装 openpyxl。运行：pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# 模板路径（相对于本脚本所在目录）
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).parent
REPO_ROOT = SCRIPT_DIR.parent

TEMPLATES = {
    "single":    REPO_ROOT / "refrences" / "公有云SA 单机模版_V3.7.2_20251125.xlsx",
    "mini":      REPO_ROOT / "refrences" / "公有云SA mini 集群模版_V3.7.2_20251125.xlsx",
    "standard":  REPO_ROOT / "refrences" / "公有云SA 标准集群模版_V3.7.2_20251125.xlsx",
    "sfn-3node": REPO_ROOT / "refrences" / "SFN 3 节点集群v2.4.4.xlsx",
    "sfn-3+3":   REPO_ROOT / "refrences" / "SFN 3+3 集群 v2.4.4.xlsx",
    "sfn-3+n+3": REPO_ROOT / "refrences" / "SFN 3+N+3 集群 v2.4.4.xlsx",
}

# ---------------------------------------------------------------------------
# 评估页填值映射
# ---------------------------------------------------------------------------

# 单机：sheet "0.评估单元"
SINGLE_EVAL_CELLS = {
    "cloud_vendor":    "C3",   # 云厂商
    "dau_range":       "C4",   # 平均日活范围
    "daily_import":    "A10",  # 日导入数据条数（亿）
    "retention_days":  "B10",  # 存储天数
    "history_events":  "C10",  # 历史已有事件量（亿）
}

# Mini 集群：sheet "0.评估页"
MINI_EVAL_CELLS = {
    "cloud_vendor":    "D3",
    "dau_range":       "D4",
    "daily_import":    "B10",
    "retention_days":  "C10",
    "history_events":  "D10",
}

# 标准集群：sheet "0.评估页"
STANDARD_EVAL_CELLS = {
    "cloud_vendor":    "D3",
    "dau_range":       "D4",
    "daily_import":    "B11",
    "retention_days":  "C11",
    "history_events":  "D11",
}

# SFN 3节点 / 3+3 / 3+N+3：sheet "0.评估单元"
SFN_EVAL_CELLS = {
    "cloud_vendor":    "C3",   # 云厂商（下拉）
    "dau_range":       "C4",   # 平均日活范围（下拉）
}

# ---------------------------------------------------------------------------
# 服务器信息页 — 列定义
# ---------------------------------------------------------------------------

# 列字母 → 字段含义
SERVER_INFO_COLS = {
    "A": "node_type",
    "B": "ip",
    "C": "hostname",
    "D": "user",
    "E": "passwd",
    "F": "port",
    "G": "system_disk_dev",
    "H": "system_disk_size",
    "I": "meta_disk_dev",
    "J": "meta_disk_size",
    "K": "seq_disk_dev",
    "L": "seq_disk_size",
    "M": "rand_disk_dev",
    "N": "rand_disk_size",
    "O": "staging_disk_dev",
    "P": "staging_disk_size",
    "Q": "remark",
}

# 部署信息区（右侧，两种集群模版结构相同）
DEPLOY_INFO_CELLS = {
    "cloud_vendor":   "P7",
    "cloud_region":   "P8",
    "ntp":            "P9",
    "timezone":       "P10",
    "login_method":   "P13",
}

# 单机模版部署信息区列不同（L/M列）
SINGLE_DEPLOY_INFO_CELLS = {
    "cloud_vendor":   "M7",
    "cloud_region":   "M8",
    "ntp":            "M9",
    "timezone":       "M10",
    "login_method":   "M13",
}

# ---------------------------------------------------------------------------
# 节点行起始位置
# ---------------------------------------------------------------------------

# mini 集群：hybrid 节点，每节点占 3 行（主行 + 2 续行）
MINI_NODE_START_ROW = 18
MINI_NODE_ROW_SPAN = 3   # 每节点占行数

# 标准集群：meta 节点每节点 1 行，data 节点每节点 3 行
STANDARD_META_START_ROW = 18
STANDARD_DATA_START_ROW = 21
STANDARD_DATA_ROW_SPAN = 3

# 单机：hybrid 节点，1 行
SINGLE_NODE_ROW = 18

# ---------------------------------------------------------------------------
# 辅助函数
# ---------------------------------------------------------------------------

def set_cell(ws, cell_addr: str, value):
    """设置单元格值，跳过合并单元格的从属格。"""
    cell = ws[cell_addr]
    # 如果是合并单元格的从属格，找到主格
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # 只写主格（左上角）
            min_col = merged_range.min_col
            min_row = merged_range.min_row
            ws.cell(row=min_row, column=min_col).value = value
            return
    cell.value = value


def fill_eval_page(wb, mode: str, params: dict):
    """填写评估页的橙色输入单元格。"""
    if mode == "single":
        sheet_name = "0.评估单元"
        cell_map = SINGLE_EVAL_CELLS
    elif mode in ("sfn-3node", "sfn-3+3", "sfn-3+n+3"):
        sheet_name = "0.评估单元"
        cell_map = SFN_EVAL_CELLS
    else:
        sheet_name = "0.评估页"
        cell_map = MINI_EVAL_CELLS if mode == "mini" else STANDARD_EVAL_CELLS

    ws = wb[sheet_name]

    if params.get("cloud_vendor"):
        set_cell(ws, cell_map["cloud_vendor"], params["cloud_vendor"])
    if params.get("dau_range"):
        set_cell(ws, cell_map["dau_range"], params["dau_range"])
    # SFN 模板评估页只有云厂商和日活两个输入项，其余字段跳过
    if mode in ("sfn-3node", "sfn-3+3", "sfn-3+n+3"):
        return
    if params.get("daily_import") is not None:
        set_cell(ws, cell_map["daily_import"], float(params["daily_import"]))
    if params.get("retention_days") is not None:
        set_cell(ws, cell_map["retention_days"], int(params["retention_days"]))
    if params.get("history_events") is not None:
        set_cell(ws, cell_map["history_events"], float(params["history_events"]))


def fill_deploy_info(wb, mode: str, params: dict):
    """填写服务器信息页右侧的部署信息区。"""
    ws = wb["3.服务器信息"]
    cell_map = SINGLE_DEPLOY_INFO_CELLS if mode == "single" else DEPLOY_INFO_CELLS

    for key, addr in cell_map.items():
        if params.get(key):
            set_cell(ws, addr, params[key])


def safe_set(ws, row: int, col: int, value):
    """写入单元格。如果目标格是合并格的从属格，直接跳过（不写入）。"""
    cell = ws.cell(row=row, column=col)
    if cell.__class__.__name__ == "MergedCell":
        return  # 从属格不可写，跳过
    cell.value = value


def write_node_row(ws, row: int, node_type: str, ip: str,
                   seq_disk_size: int, rand_disk_size: int,
                   meta_disk_size: int, staging_disk_size: int,
                   seq_disk_count: int = 3,
                   user: str = "root", passwd: str = "", port: int = 22):
    """
    写入一个节点的主行及续行（顺序盘续行）。
    seq_disk_count: 顺序盘数量（mini/data 节点通常 3 或 6 块）
    返回实际占用的行数。
    """
    # 主行
    safe_set(ws, row, 1, node_type)
    safe_set(ws, row, 2, ip)
    safe_set(ws, row, 4, user)
    safe_set(ws, row, 5, passwd)
    safe_set(ws, row, 6, port)

    # 系统盘（所有节点都有）
    safe_set(ws, row, 7, "/dev/vda1")
    safe_set(ws, row, 8, "50G")

    if node_type == "meta":
        # 元数据节点：只有系统盘 + 元数据盘
        safe_set(ws, row, 9, "/dev/vdb")
        safe_set(ws, row, 10, f"{meta_disk_size}G")
        return 1  # 占 1 行

    # data / hybrid 节点
    if node_type in ("data", "hybrid"):
        # 元数据盘（hybrid 有，data 无）
        if node_type == "hybrid":
            safe_set(ws, row, 9, "/dev/vdb")
            safe_set(ws, row, 10, f"{meta_disk_size}G")
            seq_start_dev = "c"
            rand_start_idx = ord("c") + seq_disk_count
        else:
            seq_start_dev = "b"
            rand_start_idx = ord("b") + seq_disk_count

        # 顺序盘（主行写第1块，续行写剩余）
        dev_letters = [chr(ord(seq_start_dev) + i) for i in range(seq_disk_count)]
        safe_set(ws, row, 11, f"/dev/vd{dev_letters[0]}")
        safe_set(ws, row, 12, f"{seq_disk_size}G")

        for i, letter in enumerate(dev_letters[1:], start=1):
            safe_set(ws, row + i, 11, f"/dev/vd{letter}")
            safe_set(ws, row + i, 12, f"{seq_disk_size}G")

        # 随机数据盘（第1块写在主行）
        rand_dev = chr(rand_start_idx)
        safe_set(ws, row, 13, f"/dev/vd{rand_dev}")
        safe_set(ws, row, 14, f"{rand_disk_size}G")

        # 暂存盘
        staging_dev = chr(rand_start_idx + 1)
        safe_set(ws, row, 15, f"/dev/vd{staging_dev}")
        safe_set(ws, row, 16, f"{staging_disk_size}G")

        return seq_disk_count  # 占行数 = 顺序盘数量（主行 + seq_disk_count-1 续行）

    return 1


def _fill_sfn_server_info(ws, mode: str, node_groups: list,
                          seq_disk_size: int, rand_disk_size: int,
                          meta_disk_size: int, seq_disk_count: int):
    """
    SFN 模板服务器信息页结构与 CDP 不同（两行一组格式），
    自动填写暂不支持，跳过。用户需手动填写服务器信息页。
    """
    pass  # SFN 服务器信息页结构复杂，暂不自动填写


def fill_server_info(wb, mode: str, node_groups: list,
                     seq_disk_size: int, rand_disk_size: int,
                     meta_disk_size: int, staging_disk_size: int,
                     seq_disk_count: int):
    """
    填写服务器信息页的节点行。
    node_groups: [{"type": "meta/data/hybrid/sfmix/sfonline", "ips": [...]}]
    """
    if not node_groups:
        return

    ws = wb["3.服务器信息"]

    # SFN 模板：服务器信息页结构与 CDP 相同，节点类型映射到内部类型
    # sfn-3node: 混合节点(hybrid) ×3，起始行 18，每节点 3 行
    # sfn-3+3:   元数据节点(meta) ×3 + SF混合节点(hybrid) ×3
    # sfn-3+n+3: 元数据节点(meta) ×3 + 数据节点(data) ×N + SF在线节点(sfonline) ×3
    if mode in ("sfn-3node", "sfn-3+3", "sfn-3+n+3"):
        _fill_sfn_server_info(ws, mode, node_groups,
                              seq_disk_size, rand_disk_size, meta_disk_size, seq_disk_count)
        return

    if mode == "single":
        ips = node_groups[0]["ips"] if node_groups else []
        ip = ips[0] if ips else ""
        write_node_row(ws, SINGLE_NODE_ROW, "hybrid", ip,
                       seq_disk_size, rand_disk_size, meta_disk_size, staging_disk_size,
                       seq_disk_count=1)
        return

    if mode == "mini":
        # Mini：3 个 hybrid 节点，模板每节点预留 3 行
        # 若 seq_disk_count > 3，需要插入额外行
        ips = node_groups[0]["ips"] if node_groups else []
        rows_per_node = max(seq_disk_count, 3)  # 至少 3 行（模板预留）
        template_rows_per_node = 3

        current_row = MINI_NODE_START_ROW
        for i, ip in enumerate(ips[:3]):
            # 如果需要比模板多的行，先插入
            extra = rows_per_node - template_rows_per_node
            if extra > 0:
                ws.insert_rows(current_row + template_rows_per_node, extra)
            write_node_row(ws, current_row, "hybrid", ip,
                           seq_disk_size, rand_disk_size, meta_disk_size, staging_disk_size,
                           seq_disk_count=seq_disk_count)
            current_row += rows_per_node
        return

    if mode == "standard":
        meta_ips = []
        data_ips = []
        for g in node_groups:
            if g["type"] == "meta":
                meta_ips = g["ips"]
            elif g["type"] == "data":
                data_ips = g["ips"]

        # meta 节点（每节点 1 行，直接写）
        for i, ip in enumerate(meta_ips[:3]):
            row = STANDARD_META_START_ROW + i
            write_node_row(ws, row, "meta", ip,
                           seq_disk_size, rand_disk_size, meta_disk_size, staging_disk_size)

        # data 节点：模板预留 9 行（3节点×3行），计算实际需要的行数
        # 在合并格分隔行（行30）之前插入不足的行，并解除新行的合并格状态
        data_start = STANDARD_META_START_ROW + len(meta_ips[:3])
        template_data_rows = 9          # 模板预留：3节点×3行
        needed_data_rows = len(data_ips) * seq_disk_count
        extra_rows = max(0, needed_data_rows - template_data_rows)

        separator_row = data_start + template_data_rows  # 行30
        if extra_rows > 0:
            ws.insert_rows(separator_row, extra_rows)
            # 解除插入行上可能继承的合并格状态
            to_remove = [
                mr for mr in list(ws.merged_cells.ranges)
                if separator_row <= mr.min_row <= separator_row + extra_rows - 1
            ]
            for mr in to_remove:
                ws.merged_cells.remove(mr)

        # 顺序写所有 data 节点
        current_row = data_start
        for ip in data_ips:
            write_node_row(ws, current_row, "data", ip,
                           seq_disk_size, rand_disk_size, meta_disk_size, staging_disk_size,
                           seq_disk_count=seq_disk_count)
            current_row += seq_disk_count


# ---------------------------------------------------------------------------
# 主函数
# ---------------------------------------------------------------------------

def parse_node_groups(node_args: list) -> list:
    """
    解析 --nodes 参数，格式：'meta:10.0.0.1,10.0.0.2' 或 'data:10.0.0.3,10.0.0.4'
    返回：[{"type": "meta", "ips": [...]}, ...]
    """
    groups = []
    for arg in (node_args or []):
        if ":" in arg:
            node_type, ips_str = arg.split(":", 1)
            ips = [ip.strip() for ip in ips_str.split(",") if ip.strip()]
        else:
            # 没有类型前缀，默认 hybrid
            node_type = "hybrid"
            ips = [ip.strip() for ip in arg.split(",") if ip.strip()]
        groups.append({"type": node_type.strip(), "ips": ips})
    return groups


def generate_excel(
    mode: str,
    output_path: str,
    cloud_vendor: str,
    dau_range: str,
    daily_import: float,
    retention_days: int,
    history_events: float,
    node_groups: list,
    seq_disk_size: int,
    rand_disk_size: int,
    meta_disk_size: int,
    staging_disk_size: int,
    seq_disk_count: int,
    cloud_region: str = "",
    ntp: str = "ntp.aliyun.com",
    timezone: str = "Asia/Shanghai (东八区)",
    login_method: str = "",
):
    template_path = TEMPLATES.get(mode)
    if not template_path or not template_path.exists():
        print(f"错误：找不到模板文件 {template_path}", file=sys.stderr)
        sys.exit(1)

    # 复制模板
    shutil.copy2(template_path, output_path)

    # 用 openpyxl 打开副本（keep_vba=False，keep_links=False 避免外部引用警告）
    wb = openpyxl.load_workbook(output_path, keep_vba=False)

    params = {
        "cloud_vendor":   cloud_vendor,
        "dau_range":      dau_range,
        "daily_import":   daily_import,
        "retention_days": retention_days,
        "history_events": history_events,
        "cloud_region":   cloud_region,
        "ntp":            ntp,
        "timezone":       timezone,
        "login_method":   login_method,
    }

    fill_eval_page(wb, mode, params)
    fill_deploy_info(wb, mode, params)
    fill_server_info(wb, mode, node_groups,
                     seq_disk_size, rand_disk_size, meta_disk_size, staging_disk_size,
                     seq_disk_count)

    wb.save(output_path)
    print(f"✓ 已生成：{output_path}")


def main():
    parser = argparse.ArgumentParser(description="神策服务器资源评估 Excel 生成器")
    parser.add_argument("--mode",             required=True,
                        choices=["single", "mini", "standard",
                                 "sfn-3node", "sfn-3+3", "sfn-3+n+3"],
                        help="部署模式：single/mini/standard（CDP）或 sfn-3node/sfn-3+3/sfn-3+n+3（SFN）")
    parser.add_argument("--cloud-vendor",     required=True,  help="云厂商，如 阿里云、腾讯云、AWS")
    parser.add_argument("--dau-range",        required=True,  help="日活范围，如 '<100万'、'<800万'")
    parser.add_argument("--daily-import",     type=float, default=0,
                        help="日导入数据条数（亿），如 0.5")
    parser.add_argument("--retention-days",   type=int, default=365, help="存储天数（默认 365）")
    parser.add_argument("--history-events",   type=float, default=0,
                        help="历史已有事件量（亿），如 0")
    parser.add_argument("--nodes",            action="append", default=[],
                        help="节点 IP，格式：'meta:10.0.0.1,10.0.0.2' 或 'data:10.0.0.3'，可多次指定")
    parser.add_argument("--seq-disk-size",    type=int, default=1000,
                        help="顺序数据盘单块容量 GB（默认 1000）")
    parser.add_argument("--seq-disk-count",   type=int, default=3,
                        help="每节点顺序数据盘数量（默认 3，标准集群通常 6）")
    parser.add_argument("--rand-disk-size",   type=int, default=500,
                        help="随机数据盘单块容量 GB（默认 500）")
    parser.add_argument("--meta-disk-size",   type=int, default=400,
                        help="元数据盘容量 GB（默认 400）")
    parser.add_argument("--staging-disk-size",type=int, default=250,
                        help="暂存盘容量 GB（默认 250）")
    parser.add_argument("--cloud-region",     default="",     help="云服务器所在地域，如 华北1")
    parser.add_argument("--ntp",              default="ntp.aliyun.com", help="NTP 时钟源地址")
    parser.add_argument("--timezone",         default="Asia/Shanghai (东八区)", help="时区")
    parser.add_argument("--login-method",     default="",     help="远程登录方式，如 ssh直连")
    parser.add_argument("--output",           required=True,  help="输出文件路径，如 /tmp/客户名_SA配置.xlsx")

    args = parser.parse_args()
    node_groups = parse_node_groups(args.nodes)

    generate_excel(
        mode=args.mode,
        output_path=args.output,
        cloud_vendor=args.cloud_vendor,
        dau_range=args.dau_range,
        daily_import=args.daily_import,
        retention_days=args.retention_days,
        history_events=args.history_events,
        node_groups=node_groups,
        seq_disk_size=args.seq_disk_size,
        seq_disk_count=args.seq_disk_count,
        rand_disk_size=args.rand_disk_size,
        meta_disk_size=args.meta_disk_size,
        staging_disk_size=args.staging_disk_size,
        cloud_region=args.cloud_region,
        ntp=args.ntp,
        timezone=args.timezone,
        login_method=args.login_method,
    )


if __name__ == "__main__":
    main()
